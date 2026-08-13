"""HR and compliance: the week's attendance, the gazette, the buyer's own audit."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.platypus import Spacer

import _order as O
from _lib import (
    F_SANS,
    F_SANS_B,
    P,
    Sheet,
    build_pdf,
    doc_title,
    grid,
    kv_block,
    letterhead,
    money,
    rng,
    save_scan_pdf,
    signature_row,
    two_col,
    write_json,
    write_text,
)
from d_merch import BORDER, TOT_FILL, _head, _style_sheet


# ─────────────────────────────────────────────────────────────────────────────
# 26 · the line's attendance and overtime (xlsx — payroll's input)
# ─────────────────────────────────────────────────────────────────────────────

FIRST = [
    "Rokeya", "Shahnaz", "Momena", "Rahima", "Sufia", "Jesmin", "Parvin", "Nasrin",
    "Hasina", "Ruma", "Salma", "Kohinoor", "Rehana", "Anowara", "Jorina", "Mukta",
    "Sathi", "Piyara", "Rasheda", "Bilkis", "Shefali", "Ayesha", "Rina", "Marufa",
    "Shilpi", "Taslima", "Fatema", "Nazma", "Halima", "Sabina", "Rojina", "Beauty",
    "Shapla", "Champa", "Rani", "Josna", "Lipi", "Monira", "Rahela", "Delwara",
    "Shirin", "Sanjida", "Rabeya", "Kulsum", "Amena", "Rokshana", "Nurjahan", "Sajeda",
    "Shamima", "Jahanara", "Rehena", "Farida", "Rowshan", "Momtaz", "Rina", "Runa",
    "Shahida", "Mahmuda", "Rabia", "Selina", "Papiya", "Laizu", "Sumi", "Nipa",
    "Ruksana", "Poly", "Mitu", "Konica",
]
LAST = ["Begum", "Khatun", "Akter", "Bibi", "Sultana", "Parvin"]
DESIGNATION = [
    ("Operator", "4"), ("Operator", "4"), ("Operator", "5"), ("Senior operator", "3"),
    ("Helper", "7"), ("Helper", "6"), ("Quality inspector", "4"), ("Line supervisor", "2"),
]


def attendance(out: Path) -> None:
    r = rng("attendance-l3")
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws["A1"] = f"{O.FACTORY['name']} — attendance and overtime register"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"Line {O.LINE} · week {O.ATT_WEEK[0]} to {O.ATT_WEEK[-1]} · pay period "
        f"{O.PAY_PERIOD} · style {O.STYLE} / {O.PO_NO}"
    )
    ws["A2"].font = Font(size=9, italic=True)
    ws["A3"] = (
        "Normal shift 08:00–17:00 with 1 hour unpaid lunch = 8 hours. Overtime beyond 8 hours "
        "at 2× basic hourly (basic ÷ 208). P = present, A = absent, L = leave."
    )
    ws["A3"].font = Font(size=9, italic=True)

    head = ["Card no", "Name", "Designation", "Grade"]
    for d in O.ATT_WEEK:
        head += [f"{d[5:]} in", f"{d[5:]} out", f"{d[5:]} OT"]
    head += ["Days present", "Total OT hrs"]
    _head(ws, 5, head)

    row = 6
    absent_fill = PatternFill("solid", fgColor="FBE3E0")
    for i in range(O.OPERATORS):
        desig, grade = DESIGNATION[i % len(DESIGNATION)]
        name = f"{FIRST[i % len(FIRST)]} {LAST[(i * 3) % len(LAST)]}"
        vals = [f"TT-{3400 + i:04d}", name, desig, grade]
        present = 0
        ot_total = 0.0
        for di, _d in enumerate(O.ATT_WEEK):
            # roughly 4% absence, and Tuesday is the heavy overtime day
            absent = r.random() < 0.04
            if absent:
                vals += ["A", "A", 0]
                continue
            present += 1
            ot = r.choice([0, 0, 1, 1.5, 2, 2, 2, 2.5]) if di != 4 else r.choice([0, 1, 2])
            out_h = 17 + int(ot)
            out_m = 30 if ot % 1 else 0
            vals += ["08:00", f"{out_h:02d}:{out_m:02d}", ot]
            ot_total += ot
        vals += [present, round(ot_total, 1)]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=j, value=v)
            c.border = BORDER
            if v == "A":
                c.fill = absent_fill
        row += 1

    ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
    for col in (len(head) - 1, len(head)):
        letter = chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)
        c = ws.cell(row=row, column=col, value=f"=SUM({letter}6:{letter}{row-1})")
        c.font, c.fill, c.border = Font(bold=True), TOT_FILL, BORDER

    row += 2
    ws.cell(row=row, column=1, value=(
        "NOTE FOR PAYROLL — the Bangladesh Labour Rules 2015 cap ordinary overtime at 2 hours "
        "per day. Rows showing 2.5 hours need the written consent the buyer's audit finding "
        "(major, section 5) asked for. Do not process them until HR confirms consent is on "
        "file."
    )).font = Font(italic=True, size=9, color="8A3A12")

    _style_sheet(ws, {"A": 10, "B": 20, "C": 17, "D": 7})
    ws.freeze_panes = "E6"
    p = out / "26-attendance-and-overtime-L3.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)


# ─────────────────────────────────────────────────────────────────────────────
# 27 · the wage gazette (door: wage_gazette → wage_gazettes) — a scan
# ─────────────────────────────────────────────────────────────────────────────


def gazette(out: Path) -> None:
    sh = Sheet(seed="gazette")
    W, m = sh.w, 170

    sh.bn((W // 2, 110), "গণপ্রজাতন্ত্রী বাংলাদেশ সরকার", size=46, bold=True, anchor="ma")
    sh.text((W // 2, 180), "GOVERNMENT OF THE PEOPLE'S REPUBLIC OF BANGLADESH", F_SANS_B,
            30, anchor="ma")
    sh.bn((W // 2, 232), "শ্রম ও কর্মসংস্থান মন্ত্রণালয়", size=34, anchor="ma")
    sh.text((W // 2, 292), "MINISTRY OF LABOUR AND EMPLOYMENT", F_SANS, 26,
            fill=(70, 70, 70), anchor="ma")
    sh.text((W // 2, 336), "Minimum Wages Board", F_SANS, 26, fill=(70, 70, 70), anchor="ma")
    sh.hrule(386, m, W - m, w=4)

    sh.text((W // 2, 410), "NOTIFICATION", F_SANS_B, 34, anchor="ma")
    sh.text((m, 470), f"Dhaka, the 24th November 2026", F_SANS, 26)
    sh.text((W - m, 470), f"No. {O.GAZETTE_VERSION}", F_SANS_B, 30, anchor="ra")

    y = 540
    for ln in [
        "In exercise of the powers conferred by section 140(1) of the Bangladesh Labour Act, 2006 "
        "(Act No. XLII of",
        "2006), and on the recommendation of the Minimum Wages Board, the Government is pleased to "
        "declare the",
        "following minimum rates of wages for workers employed in the ready-made garment sector, "
        "with effect from",
        f"{O.GAZETTE_FROM}. These rates supersede all previous notifications for this sector.",
    ]:
        sh.text((m, y), ln, F_SANS, 26, fill=(35, 35, 35))
        y += 44

    y += 40
    sh.text((m, y), "SCHEDULE — MONTHLY MINIMUM WAGE BY GRADE (in Taka)", F_SANS_B, 30)
    y += 56
    rows = [["Grade", "Basic", "House rent", "Medical", "Transport", "Food", "Gross"]]
    for g, basic, house, med, trans, food in O.GAZETTE_GRADES:
        gross = int(basic) + int(house) + int(med) + int(trans) + int(food)
        rows.append([g, f"{int(basic):,}", f"{int(house):,}", f"{int(med):,}",
                     f"{int(trans):,}", f"{int(food):,}", f"{gross:,}"])
    y = sh.table(m, y, [220, 320, 330, 280, 320, 280, 390], rows, row_h=78, size=28,
                 align=["c", "r", "r", "r", "r", "r", "r"])

    y += 46
    for ln in [
        "1. House rent is 50 per cent of the basic wage. Medical, transport and food allowances are "
        "as stated and",
        "   are not subject to grade.",
        "2. Overtime shall be paid at twice the ordinary hourly rate, the ordinary hourly rate being "
        "the basic wage",
        "   divided by 208.",
        "3. Two festival bonuses shall be paid in each calendar year, each equal to one month's basic "
        "wage, pro-rated",
        "   for workers with less than one year of continuous service.",
        "4. No worker shall be paid less than the rate of the grade in which the worker is classified.",
    ]:
        sh.text((m, y), ln, F_SANS, 25, fill=(45, 45, 45))
        y += 42

    y += 50
    sh.text((W - m, y), "By order of the President", F_SANS, 26, anchor="ra")
    sh.sign((W - m - 420, y + 70), "gazette-sec", scale=0.8, color=(20, 30, 80))
    sh.line((W - m - 520, y + 158), (W - m, y + 158), fill=(90, 90, 90), w=2)
    sh.text((W - m, y + 174), "Secretary, Minimum Wages Board", F_SANS, 24,
            fill=(70, 70, 70), anchor="ra")
    sh.stamp((m + 300, y - 20), ["MINIMUM", "WAGES BOARD", "DHAKA"], r=160, rot=-8,
             color=(40, 70, 40))

    sh.text((m, sh.h - 130), "TEST FIXTURE — generated for FabricXAI platform testing. "
            "Not a real gazette notification and carries no legal force.", F_SANS, 22,
            fill=(140, 140, 140))

    save_scan_pdf(sh, out / "27-wage-gazette-SRO-2026-11.pdf", "gazette-scan", grain=14,
                  dark=0.945)

    grades_txt = "\n".join(
        f"{g:<8}{int(b):>10,}{int(h):>13,}{int(md):>11,}{int(t):>12,}{int(f):>9,}"
        f"{int(b)+int(h)+int(md)+int(t)+int(f):>11,}"
        for g, b, h, md, t, f in O.GAZETTE_GRADES
    )
    write_text(
        out / "27-wage-gazette-SRO-2026-11.paste.txt",
        f"""GOVERNMENT OF THE PEOPLE'S REPUBLIC OF BANGLADESH
MINISTRY OF LABOUR AND EMPLOYMENT — Minimum Wages Board

NOTIFICATION
Dhaka, the 24th November 2026                                        No. {O.GAZETTE_VERSION}

In exercise of the powers conferred by section 140(1) of the Bangladesh Labour Act, 2006
(Act No. XLII of 2006), and on the recommendation of the Minimum Wages Board, the
Government is pleased to declare the following minimum rates of wages for workers employed
in the ready-made garment sector, with effect from {O.GAZETTE_FROM}. These rates supersede
all previous notifications for this sector.

SCHEDULE — MONTHLY MINIMUM WAGE BY GRADE (in Taka)
Grade        Basic   House rent    Medical   Transport     Food      Gross
{grades_txt}

1. House rent is 50 per cent of the basic wage.
2. Overtime shall be paid at twice the ordinary hourly rate, the ordinary hourly rate being
   the basic wage divided by 208.
3. Two festival bonuses shall be paid in each calendar year, each equal to one month's basic
   wage, pro-rated for workers with less than one year of continuous service.
""",
    )
    write_json(
        out / "27-wage-gazette-SRO-2026-11.expected.json",
        {
            "_intakeKind": "wage_gazette",
            "_door": "/marbim/intake → 'A wage gazette notification' (hr only)",
            "version": O.GAZETTE_VERSION,
            "effectiveFrom": O.GAZETTE_FROM,
            "grades": [
                {"grade": g, "basic": b, "houseRent": h, "medical": md,
                 "transport": t, "food": f}
                for g, b, h, md, t, f in O.GAZETTE_GRADES
            ],
            "_notes": [
                "TENANT-WIDE, not order-specific. Only run this if the tenant does not already "
                "carry an active gazette — approving a second one changes every payroll on the "
                "tenant, including the other kit's. It lands INACTIVE and must be activated by "
                "hand; that two-step is the thing to verify.",
                "Seven grades. The Gross column is the SUM of the other five and is not a "
                "field — a gazette whose basic comes back as 17,250 has read the gross.",
                "Amounts are whole taka as strings: '10400', not '10,400' and not 10400.00.",
                "Role wall: only hr (and owner/admin) may file this. A merchandiser who can "
                "reach the chip is a finding.",
            ],
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 28 · the buyer's own audit report (door: audit_report → findings)
# ─────────────────────────────────────────────────────────────────────────────

SEV_LABEL = {"critical": "Critical (zero tolerance)", "major": "Major",
             "minor": "Minor", "observation": "Observation"}


def audit_report(out: Path) -> None:
    counts = {s: sum(1 for f in O.FINDINGS if f["sev"] == s)
              for s in ("critical", "major", "minor", "observation")}
    rows = [["#", "Severity", "Finding", "Ref"]]
    for i, f in enumerate(O.FINDINGS, 1):
        rows.append([str(i), SEV_LABEL[f["sev"]], P(f["text"], "small"), f"p.{f['page']}"])

    st = [
        *letterhead(
            O.AUDIT_FIRM,
            ["Social compliance and supply chain assurance",
             "House 22, Road 113, Gulshan 2, Dhaka 1212, Bangladesh"],
            [f"Report {O.AUDIT_REF}", f"Issued {O.AUDIT_DATE}", "Accreditation VA/BD/2019/044"],
            accent="#4a3410",
        ),
        *doc_title(
            "Supplier social compliance audit — report",
            f"On behalf of {O.BUYER['name']} · against {O.AUDIT_STANDARD}",
        ),
        two_col(
            [
                kv_block(
                    [
                        ("Facility", O.FACTORY["name"]),
                        ("Address", O.FACTORY["addr"][0] + ", " + O.FACTORY["addr"][1]),
                        ("Audit type", "Initial — pre-order, announced"),
                        ("Audit date", O.AUDIT_DATE),
                        ("Auditor", "S. Chowdhury, lead · badge VA-BD-0918"),
                    ],
                    widths=(28, 58),
                )
            ],
            [
                kv_block(
                    [
                        ("Standard", O.AUDIT_STANDARD),
                        ("Workers on site", "1,284 (947 female, 337 male)"),
                        ("Workers interviewed", "38 (28 individual, 2 groups)"),
                        ("Records period reviewed", "June – July 2026"),
                        ("Overall rating", "C — conditional, re-audit in 90 days"),
                    ],
                    widths=(30, 56),
                )
            ],
        ),
        Spacer(1, 8),
        grid(
            [
                ["Critical", "Major", "Minor", "Observation", "Total"],
                [str(counts["critical"]), str(counts["major"]), str(counts["minor"]),
                 str(counts["observation"]), str(len(O.FINDINGS))],
            ],
            [36, 36, 36, 40, 32],
            align_right=[0, 1, 2, 3, 4],
        ),
        Spacer(1, 4),
        P(
            "A critical finding suspends new order placement until it is closed and verified. "
            f"{O.BUYER['name']} has placed {O.PO_NO} conditionally on the fire-exit finding "
            "being closed within 14 days with photographic evidence.",
            "small",
        ),
        Spacer(1, 9),
        P("Findings", "h2"),
        grid(rows, [10, 34, 122, 14], font_size=7.6),
        Spacer(1, 9),
        P("Corrective action plan", "h2"),
        grid(
            [["#", "Action required", "Due", "Owner"]]
            + [
                [str(i), P(action, "small"), due, owner]
                for i, (action, due, owner) in enumerate(
                    [
                        ("Clear and unlock all fire exits; move carton staging out of the "
                         "egress path; keys removed from the security office.",
                         "2026-09-02", "Compliance"),
                        ("Obtain and file written overtime consent; cap daily OT at 2 hours "
                         "in the attendance system.", "2026-09-30", "HR"),
                        ("Install bunded storage and display Bangla MSDS; label all "
                         "containers.", "2026-09-30", "Compliance"),
                        ("Re-constitute the anti-harassment committee and minute monthly "
                         "meetings.", "2026-10-15", "HR"),
                        ("Add drinking-water points to reach 1 per 50 workers.",
                         "2026-10-15", "Admin"),
                        ("Re-train cutters on PPE; daily supervisor check.",
                         "2026-09-15", "Quality"),
                        ("Fit guards to all button-attach machines.",
                         "2026-09-15", "Maintenance"),
                    ],
                    start=1,
                )
            ],
            [10, 116, 30, 24],
            font_size=7.6,
        ),
        Spacer(1, 8),
        signature_row(["Lead auditor · " + O.AUDIT_FIRM, "Facility representative"]),
    ]
    build_pdf(out / "28-buyer-audit-report.pdf", st, "Supplier compliance audit report")

    findings_txt = "\n\n".join(
        f"{i}. [{SEV_LABEL[f['sev']]}] (p.{f['page']})\n   {f['text']}"
        for i, f in enumerate(O.FINDINGS, 1)
    )
    write_text(
        out / "28-buyer-audit-report.paste.txt",
        f"""{O.AUDIT_FIRM}
Social compliance and supply chain assurance

SUPPLIER SOCIAL COMPLIANCE AUDIT — REPORT
On behalf of {O.BUYER['name']} · against {O.AUDIT_STANDARD}

Report: {O.AUDIT_REF}
Issued: {O.AUDIT_DATE}
Facility: {O.FACTORY['name']}, {O.FACTORY['addr'][0]}, {O.FACTORY['addr'][1]}
Audit type: Initial — pre-order, announced
Audit date: {O.AUDIT_DATE}
Auditor: S. Chowdhury, lead · badge VA-BD-0918
Workers on site: 1,284 (947 female, 337 male)
Records period reviewed: June – July 2026
Overall rating: C — conditional, re-audit in 90 days

Critical: {counts['critical']}   Major: {counts['major']}   Minor: {counts['minor']}   """
        f"""Observation: {counts['observation']}   Total: {len(O.FINDINGS)}

FINDINGS

{findings_txt}
""",
    )
    write_json(
        out / "28-buyer-audit-report.expected.json",
        {
            "_intakeKind": "audit_report",
            "_door": "/marbim/intake → 'A compliance audit report' (compliance only)",
            "_context": {
                "auditId": "picked from the Audit dropdown. Create the audit FIRST — "
                f"regime 'buyer', {O.AUDIT_DATE}, {O.AUDIT_FIRM} — then file this against it."
            },
            "findings": [
                {"severity": f["sev"], "text": f["text"], "sourcePage": f["page"],
                 "evidence": []}
                for f in O.FINDINGS
            ],
            "_notes": [
                "EIGHT findings: 1 critical, 3 major, 3 minor, 1 observation. The severity "
                "words on the page map onto the enum exactly — 'Critical (zero tolerance)' is "
                "still `critical`.",
                "The corrective action plan at the end is NOT a finding list. Seven more "
                "findings that read like instructions ('Clear and unlock all fire exits') "
                "means the CAP table was swallowed.",
                "The critical finding is the one with teeth: it should gate the order in the "
                "compliance screen, and the kit's walk checks that it does.",
            ],
        },
    )


def build(out: Path) -> None:
    attendance(out)
    gazette(out)
    audit_report(out)
