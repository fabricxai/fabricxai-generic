"""Quality: the chart it measures against, the fabric it rejected, the lot it passed."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.platypus import Spacer

import _order as O
from _lib import (
    P,
    build_pdf,
    doc_title,
    grid,
    kv_block,
    letterhead,
    money,
    plain,
    signature_row,
    two_col,
    write_json,
    write_text,
)
from d_merch import BORDER, HEAD_FILL, TOT_FILL, BUYER_HEAD, FACTORY_HEAD, _head, _style_sheet


# ─────────────────────────────────────────────────────────────────────────────
# 15 · the measurement chart (door: measurement_chart → measurement_specs)
# ─────────────────────────────────────────────────────────────────────────────


def _pom_rows() -> list[list]:
    rows = [["POM", "Point of measure", *O.SIZES, "Tol ±"]]
    for p in O.POM:
        rows.append(
            [p["code"], p["name"]] + [plain(v, 1) for v in p["v"]] + [plain(p["tol"], 1)]
        )
    return rows


def measurement_chart(out: Path) -> None:
    st = [
        *BUYER_HEAD(),
        *doc_title(
            "Measurement chart — points of measure",
            f"{O.STYLE} · {O.BUYER_ARTICLE} · {O.TECHPACK_REV} · all measurements in "
            f"{O.POM_UNIT}, garment measured flat",
        ),
        two_col(
            [kv_block([("Style", O.STYLE), ("Description", O.STYLE_DESC),
                       ("Sample stage", "PP and bulk")], widths=(28, 58))],
            [kv_block([("Issued", O.TECHPACK_DATE), ("Unit", O.POM_UNIT),
                       ("Grade rule", "as printed per size — do not interpolate")],
                      widths=(28, 58))],
        ),
        Spacer(1, 9),
        grid(_pom_rows(), [14, 76, 16, 16, 16, 16, 16, 16], align_right=[2, 3, 4, 5, 6, 7]),
        Spacer(1, 9),
        P("How these are taken", "h2"),
        P(
            "A — lay flat, measure 1 cm below the armhole seam, edge to edge.<br/>"
            "B — from the high point of shoulder straight down to the bottom edge of the rib.<br/>"
            "C — seam to seam across the back.<br/>"
            "D — from the centre back neck along the shoulder and down the sleeve to the "
            "cuff edge.<br/>"
            "E — straight from the shoulder point to the underarm point.<br/>"
            "F, G — relaxed, not stretched. Allow the rib to settle 10 seconds before "
            "measuring.<br/>"
            "H — from the top of the hood seam to the bottom of the hood opening.<br/>"
            "J — along the pocket opening edge, one side.",
            "p",
        ),
        Spacer(1, 6),
        P(
            "Tolerance is symmetric unless stated otherwise. A measurement outside tolerance "
            "on any graded point is a major defect at final inspection. Measure a minimum of "
            "3 pieces per size per colour at inline and 8 pieces per size at final.",
            "small",
        ),
    ]
    build_pdf(out / "15-measurement-chart-ST-2815.pdf", st, f"Measurement chart {O.STYLE}")

    # spreadsheet — the form QC actually keeps
    wb = Workbook()
    ws = wb.active
    ws.title = "POM"
    ws["A1"] = f"{O.BUYER['name']} — measurement chart {O.STYLE} ({O.TECHPACK_REV})"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"All measurements in {O.POM_UNIT}, garment measured flat"
    ws["A2"].font = Font(size=9, italic=True)
    _head(ws, 4, ["POM", "Point of measure", *O.SIZES, "Tol +", "Tol −"])
    r = 5
    for p in O.POM:
        vals = [p["code"], p["name"], *p["v"], p["tol"], p["tol"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            if j >= 3:
                c.number_format = "0.0"
        r += 1
    _style_sheet(ws, {"A": 6, "B": 38, "C": 8, "D": 8, "E": 8, "F": 8, "G": 8, "H": 8, "I": 8})
    ws.freeze_panes = "C5"
    p = out / "15-measurement-chart-ST-2815.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)

    pom_txt = "\n".join(
        f"{p['code']}  {p['name']:<42}" + "".join(f"{plain(v, 1):>8}" for v in p["v"])
        + f"{plain(p['tol'], 1):>9}"
        for p in O.POM
    )
    write_text(
        out / "15-measurement-chart-ST-2815.paste.txt",
        f"""{O.BUYER['name']}
MEASUREMENT CHART — POINTS OF MEASURE
{O.STYLE} · {O.BUYER_ARTICLE} · {O.TECHPACK_REV}
All measurements in {O.POM_UNIT}, garment measured flat.

POM Point of measure                                {'  '.join(f'{s:>5}' for s in O.SIZES)}    Tol +/-
{pom_txt}

Tolerance is symmetric unless stated otherwise. A measurement outside tolerance on any
graded point is a major defect at final inspection.
""",
    )
    write_json(
        out / "15-measurement-chart-ST-2815.expected.json",
        {
            "_intakeKind": "measurement_chart",
            "_door": "/marbim/intake → 'A measurement chart'",
            "styleCode": O.STYLE,
            "unit": O.POM_UNIT,
            "_pointCount": len(O.POM) * len(O.SIZES),
            "points": [
                {
                    "name": f"{p['code']} {p['name']} — size {s}",
                    "spec": plain(p["v"][i], 1),
                    "tolPlus": plain(p["tol"], 1),
                    "tolMinus": plain(p["tol"], 1),
                }
                for p in O.POM
                for i, s in enumerate(O.SIZES)
            ],
            "_notes": [
                "FIFTY points, not ten. A graded row is one point PER SIZE — the schema says "
                "so in its own description. Ten points with five values each cannot be stored "
                "and will fail, or worse, keep only the first size.",
                "The chart prints ONE 'Tol ±' column. The schema folds a single stated "
                "magnitude across both directions, so tolPlus and tolMinus should both come "
                "back as that number. Both absent means the fold did not run.",
                "Negative test worth running: paste this file as ONE line (join every row "
                "with spaces) and extract again. Columns pair with the wrong size at high "
                "confidence — a real historical failure. The approve inbox is where a human "
                "is supposed to catch it, so check that it is catchable.",
                "Row F (cuff width) repeats 8.5 for XS and S and 9.0 for M and L. Repeated "
                "values across adjacent sizes are where an off-by-one column read hides.",
            ],
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 16 · the 4-point fabric inspection (no door — quality types it)
# ─────────────────────────────────────────────────────────────────────────────

INSPECTED = [
    # roll, yards inspected, points, verdict
    ("R-F-03", 128, 9, "pass"),
    ("R-F-08", 131, 14, "pass"),
    ("R-F-12", 126, 6, "pass"),
    ("R-F-17", 129, 24, "FAIL"),
    ("R-F-21", 134, 11, "pass"),
    ("R-F-26", 127, 17, "pass"),
    ("R-F-33", 130, 8, "pass"),
    ("R-F-39", 128, 13, "pass"),
    ("R-F-44", 132, 27, "FAIL"),
    ("R-F-49", 129, 12, "pass"),
    ("R-F-54", 126, 15, "pass"),
    ("R-F-58", 131, 22, "FAIL"),
]


def four_point(out: Path) -> None:
    rows = [["Roll", "Shade", "Yards inspected", "Defect points", "Points / 100 sq yd", "Verdict"]]
    for roll, yards, pts, verdict in INSPECTED:
        shade = "A" if int(roll.split("-")[-1]) < O.GRN_SHADE_B_FROM else "B"
        per100 = round(pts * 100 / (yards * 185 / 91.44), 1)
        rows.append([roll, shade, str(yards), str(pts), plain(per100, 1), verdict])

    passed = sum(1 for *_x, v in INSPECTED if v == "pass")
    st = [
        *FACTORY_HEAD(),
        *doc_title(
            "Fabric inspection report — 4-point system",
            f"ASTM D5430 · lot {O.GRN_LOT} · challan {O.GRN_CHALLAN}",
        ),
        two_col(
            [
                kv_block(
                    [
                        ("Report no", "FI-FAB-2815-01"),
                        ("Inspection date", "2026-11-13"),
                        ("Material", "FAB-FLC-280 brushed fleece 280 g/m²"),
                        ("Supplier", O.MILL["name"]),
                        ("Colour", "Charcoal Melange"),
                    ],
                    widths=(30, 56),
                )
            ],
            [
                kv_block(
                    [
                        ("Lot received", f"{O.GRN_ROLLS} rolls · {O.GRN_KG:,.1f} kg"),
                        ("Rolls inspected", f"{len(INSPECTED)} ({len(INSPECTED)*100//O.GRN_ROLLS}%)"),
                        ("Acceptance limit", "20 points per 100 sq yd"),
                        ("Width", "185 cm (72.8 in) tubular"),
                        ("Bonded", f"yes — UD {O.UD_NO}"),
                    ],
                    widths=(30, 56),
                )
            ],
        ),
        Spacer(1, 9),
        grid(rows, [24, 18, 32, 30, 40, 26], align_right=[2, 3, 4]),
        Spacer(1, 9),
        two_col(
            [
                P("Result", "h2"),
                grid(
                    [
                        ["", "Rolls"],
                        ["Inspected", str(len(INSPECTED))],
                        ["Passed", str(passed)],
                        ["Failed", str(len(INSPECTED) - passed)],
                        ["Lot verdict", "ACCEPTED WITH SEGREGATION"],
                    ],
                    [52, 26],
                    align_right=[1],
                ),
            ],
            [
                P("Defects recorded", "h2"),
                P(
                    "R-F-17 — running slub 3.2 m from head, needle line 1.8 m; 24 pts.<br/>"
                    "R-F-44 — oil stain patch 40 × 12 cm, hole at 68 m; 27 pts.<br/>"
                    "R-F-58 — barré across full width in three places; 22 pts.<br/><br/>"
                    "Shade continuity checked against the approved swatch under D65. "
                    "R-F-39 onward reads half a step darker and is grouped B.",
                    "p",
                ),
            ],
        ),
        Spacer(1, 8),
        P(
            f"<b>Decision.</b> The lot is accepted for the passing rolls. "
            f"{', '.join(O.GRN_FAILED)} are quarantined for claim against the mill under "
            "clause 4 of the proforma and must not be issued to cutting. Shade groups A and "
            "B are to be issued to separate lays; no garment may carry both.",
            "p",
        ),
        Spacer(1, 8),
        signature_row(["Fabric QC inspector", "Quality manager", "Store — segregation confirmed"]),
    ]
    build_pdf(out / "16-fabric-4point-inspection.pdf", st, "4-point fabric inspection")

    write_text(
        out / "16-fabric-4point-inspection.paste.txt",
        f"""{O.FACTORY['name']}
FABRIC INSPECTION REPORT — 4-POINT SYSTEM (ASTM D5430)
Lot {O.GRN_LOT} · challan {O.GRN_CHALLAN}

Report no: FI-FAB-2815-01
Inspection date: 2026-11-13
Material: FAB-FLC-280 brushed fleece 280 g/m2
Supplier: {O.MILL['name']}
Colour: Charcoal Melange
Lot received: {O.GRN_ROLLS} rolls · {O.GRN_KG:,.1f} kg
Rolls inspected: {len(INSPECTED)} ({len(INSPECTED)*100//O.GRN_ROLLS}%)
Acceptance limit: 20 points per 100 sq yd

Roll     Shade  Yards inspected  Defect points  Points/100 sq yd  Verdict
"""
        + "\n".join(
            f"{roll:<9}{'A' if int(roll.split('-')[-1]) < O.GRN_SHADE_B_FROM else 'B':<7}"
            f"{yards:>15}{pts:>15}{plain(round(pts * 100 / (yards * 185 / 91.44), 1), 1):>18}  {verdict}"
            for roll, yards, pts, verdict in INSPECTED
        )
        + f"""

RESULT
Inspected: {len(INSPECTED)} rolls · Passed: {passed} · Failed: {len(INSPECTED) - passed}
Lot verdict: ACCEPTED WITH SEGREGATION

{', '.join(O.GRN_FAILED)} are quarantined for claim against the mill and must not be issued
to cutting. Shade groups A and B are to be issued to separate lays.
""",
    )


# ─────────────────────────────────────────────────────────────────────────────
# 17 · the inline tally (xlsx — the QC walk's own sheet)
# ─────────────────────────────────────────────────────────────────────────────


def inline_tally(out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inline QC"
    ws["A1"] = f"{O.FACTORY['name']} — inline QC hourly tally"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"Line {O.LINE} · {O.HOURLY_DATE} · style {O.STYLE} · {O.PO_NO} · "
        f"checked {O.HOURLY_TOTAL:,} pcs · {O.DEFECT_TOTAL} defects"
    )
    ws["A2"].font = Font(size=9, italic=True)

    hours = [h for h, _t, _a, _r in O.HOURS]
    _head(ws, 4, ["Defect code", "Defect", *[f"{h}:00" for h in hours], "Total"])
    r = 5
    # spread each defect's total across the hours deterministically
    for code, name, total in O.DEFECTS:
        per = [total // len(hours)] * len(hours)
        for i in range(total - sum(per)):
            per[(i * 3 + len(code)) % len(hours)] += 1
        vals = [code, name, *per, total]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
        r += 1
    ws.cell(row=r, column=2, value="Defects found").font = Font(bold=True)
    for i, h in enumerate(hours):
        col = 3 + i
        c = ws.cell(row=r, column=col,
                    value=f"=SUM({chr(64+col)}5:{chr(64+col)}{r-1})")
        c.font, c.fill, c.border = Font(bold=True), TOT_FILL, BORDER
    c = ws.cell(row=r, column=3 + len(hours), value=O.DEFECT_TOTAL)
    c.font, c.fill, c.border = Font(bold=True), TOT_FILL, BORDER

    r += 1
    ws.cell(row=r, column=2, value="Pieces checked").font = Font(bold=True)
    for i, (_h, _t, actual, _rm) in enumerate(O.HOURS):
        c = ws.cell(row=r, column=3 + i, value=actual)
        c.border = BORDER
    c = ws.cell(row=r, column=3 + len(hours), value=O.HOURLY_TOTAL)
    c.font, c.border = Font(bold=True), BORDER

    r += 1
    ws.cell(row=r, column=2, value="DHU %").font = Font(bold=True)
    dhu = round(O.DEFECT_TOTAL * 100 / O.HOURLY_TOTAL, 2)
    for i in range(len(hours)):
        col = 3 + i
        c = ws.cell(row=r, column=col,
                    value=f"={chr(64+col)}{r-2}/{chr(64+col)}{r-1}*100")
        c.number_format = "0.00"
        c.border = BORDER
    c = ws.cell(row=r, column=3 + len(hours), value=dhu)
    c.font, c.number_format, c.border = Font(bold=True), "0.00", BORDER

    r += 3
    ws.cell(row=r, column=1, value=(
        f"DHU for the day = {O.DEFECT_TOTAL} defects / {O.HOURLY_TOTAL:,} pieces checked = "
        f"{dhu}%. Alert threshold for this style is 5.00% — the line is above it and the "
        "supervisor was asked for a corrective action on broken and skipped stitch, which "
        "together are 45% of the day's defects."
    )).font = Font(italic=True, size=9, color="8A3A12")

    _style_sheet(ws, {"A": 18, "B": 30, **{chr(67 + i): 8 for i in range(len(hours) + 1)}})
    ws.freeze_panes = "C5"
    p = out / "17-inline-qc-dhu-tally.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)


# ─────────────────────────────────────────────────────────────────────────────
# 18 · the final inspection (no door — quality types the verdict)
# ─────────────────────────────────────────────────────────────────────────────


def final_inspection(out: Path) -> None:
    st = [
        *letterhead(
            "Nordkap Apparel AB — nominated inspection",
            ["Inspection carried out at " + O.FACTORY["name"] + ", Ashulia",
             "By Vertas Assurance Ltd on behalf of the buyer"],
            ["Report " + O.AQL_REPORT, "Date " + O.AQL_DATE, "Inspector badge VA-BD-1174"],
            accent=O.BUYER["accent"],
        ),
        *doc_title("Final random inspection report", f"{O.STYLE} · {O.PO_NO} · shipment 1 of 3"),
        two_col(
            [
                kv_block(
                    [
                        ("Style", f"{O.STYLE} / {O.BUYER_ARTICLE}"),
                        ("PO", O.PO_NO),
                        ("Colour", "Charcoal Melange"),
                        ("Offered quantity", f"{O.AQL_LOT:,} pcs"),
                        ("Cartons offered", f"{O.SHIP1_CARTONS:,}"),
                        ("Packed", f"{O.PCS_PER_CARTON} pcs solid size solid colour"),
                    ],
                    widths=(30, 56),
                )
            ],
            [
                kv_block(
                    [
                        ("Standard", "ISO 2859-1, single sampling, normal"),
                        ("Inspection level", O.AQL_LEVEL),
                        ("Sample size code", "M"),
                        ("Sample size", f"{O.AQL_SAMPLE} pcs"),
                        ("AQL", f"{O.AQL_MAJOR} major / {O.AQL_MINOR} minor"),
                        ("Cartons drawn", "23, at random across the pallet"),
                    ],
                    widths=(30, 56),
                )
            ],
        ),
        Spacer(1, 9),
        grid(
            [
                ["Class", "AQL", "Accept", "Reject", "Found", "Result"],
                ["Critical", "0", "0", "1", "0", "PASS"],
                ["Major", plain(O.AQL_MAJOR, 1), str(O.AQL_ACC_MAJOR), str(O.AQL_REJ_MAJOR),
                 str(O.AQL_FOUND_MAJOR), "PASS"],
                ["Minor", plain(O.AQL_MINOR, 1), str(O.AQL_ACC_MINOR), str(O.AQL_REJ_MINOR),
                 str(O.AQL_FOUND_MINOR), "PASS"],
            ],
            [30, 22, 26, 26, 26, 50],
            align_right=[1, 2, 3, 4],
        ),
        Spacer(1, 9),
        two_col(
            [
                P("Defects found", "h2"),
                grid(
                    [
                        ["Defect", "Maj", "Min"],
                        ["Broken / skipped stitch", "3", "5"],
                        ["Puckering at zip tape", "2", "4"],
                        ["Measurement out of tolerance (B, G)", "3", "2"],
                        ["Oil stain, washable", "0", "4"],
                        ["Loose thread / untrimmed", "0", "3"],
                        ["Zipper not running free", "1", "0"],
                        ["Total", str(O.AQL_FOUND_MAJOR), str(O.AQL_FOUND_MINOR)],
                    ],
                    [58, 14, 14],
                    align_right=[1, 2],
                ),
            ],
            [
                P("On-site tests", "h2"),
                grid(
                    [
                        ["Test", "Result"],
                        ["Measurement, 8 pcs per size", "within tolerance"],
                        ["Zipper pull test 3 kg × 10 s", "pass"],
                        ["Needle / metal detection", "100% passed, log seen"],
                        ["Carton drop test, 1 carton", "pass"],
                        ["Carton weight vs packing list", "±0.2 kg, pass"],
                        ["Barcode scan, 20 polybags", "pass"],
                        ["Shade against approved swatch", "pass — group A only"],
                    ],
                    [50, 36],
                ),
            ],
        ),
        Spacer(1, 8),
        P(
            f"<b>Overall result: PASSED.</b> The lot of {O.AQL_LOT:,} pcs is approved for "
            "shipment. Note for the vendor: measurement defects were all on points B and G "
            "in size L; the rib is settling short after the wash. Correct before shipment 2. "
            "Three cartons were re-taped after opening; carton numbers noted on the packing "
            "list copy held by the shipping desk.",
            "p",
        ),
        Spacer(1, 8),
        signature_row(["Inspector · Vertas Assurance Ltd", "Factory QA representative"]),
    ]
    build_pdf(out / "18-final-inspection-AQL-report.pdf", st, "Final inspection report")

    write_text(
        out / "18-final-inspection-AQL-report.paste.txt",
        f"""NORDKAP APPAREL AB — NOMINATED INSPECTION
FINAL RANDOM INSPECTION REPORT
{O.STYLE} · {O.PO_NO} · shipment 1 of 3

Report: {O.AQL_REPORT}
Date: {O.AQL_DATE}
Offered quantity: {O.AQL_LOT:,} pcs in {O.SHIP1_CARTONS:,} cartons
Standard: ISO 2859-1, single sampling, normal
Inspection level: {O.AQL_LEVEL}   Sample size code: M   Sample size: {O.AQL_SAMPLE} pcs
AQL: {O.AQL_MAJOR} major / {O.AQL_MINOR} minor

Class     AQL   Accept  Reject  Found  Result
Critical  0     0       1       0      PASS
Major     {O.AQL_MAJOR}   {O.AQL_ACC_MAJOR}      {O.AQL_REJ_MAJOR}      {O.AQL_FOUND_MAJOR}      PASS
Minor     {O.AQL_MINOR}   {O.AQL_ACC_MINOR}      {O.AQL_REJ_MINOR}      {O.AQL_FOUND_MINOR}     PASS

OVERALL RESULT: PASSED. The lot of {O.AQL_LOT:,} pcs is approved for shipment.
""",
    )


def build(out: Path) -> None:
    measurement_chart(out)
    four_point(out)
    inline_tally(out)
    final_inspection(out)
