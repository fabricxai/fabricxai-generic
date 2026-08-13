"""Shipment: the pack-out sheet, the full carton list, the invoice and the EXP."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.platypus import Spacer

import _order as O
from _lib import (
    F_SANS,
    F_SANS_B,
    P,
    Sheet,
    build_pdf,
    doc_title,
    grid,
    kv_block,
    letterhead,
    money,
    plain,
    save_scan_pdf,
    signature_row,
    two_col,
    write_json,
    write_text,
)
from d_merch import BORDER, TOT_FILL, FACTORY_HEAD, _head, _style_sheet


def cartons(n_from: int, n_to: int) -> list[dict]:
    """
    Shipment 1 is Charcoal Melange only, packed solid size solid colour, 24 to a carton.

    The size mix follows the PO's charcoal grid so the carton list and the order agree:
    XS 1800 / S 3600 / M 5400 / L 3600 / XL 1800, first 12,000 pieces of it.
    """
    plan: list[tuple[str, int]] = []
    remaining = O.SHIP1_QTY
    grid_ = O.BREAKDOWN["Charcoal Melange"]
    for s in O.SIZES:
        take = min(grid_[s], remaining)
        take -= take % O.PCS_PER_CARTON
        plan += [(s, O.PCS_PER_CARTON)] * (take // O.PCS_PER_CARTON)
        remaining -= take
    out = []
    for i in range(n_from, min(n_to, len(plan)) + 1):
        size, qty = plan[i - 1]
        out.append(
            dict(no=f"NKA-{i:04d}", color="Charcoal Melange", size=size, qty=qty,
                 net=O.CTN_NET, gross=O.CTN_GROSS)
        )
    return out


ALL_CARTONS = cartons(1, 10_000)


# ─────────────────────────────────────────────────────────────────────────────
# 22 · the pack-out sheet the desk types in (door: packing_list → cartons)
# ─────────────────────────────────────────────────────────────────────────────


def packout_sheet(out: Path) -> None:
    sub = ALL_CARTONS[: O.PACKOUT_CARTONS]
    rows = [["Carton no", "Colour", "Size", "Qty", "Net kg", "Gross kg"]]
    for c in sub:
        rows.append([c["no"], c["color"], c["size"], str(c["qty"]),
                     plain(c["net"]), plain(c["gross"])])
    rows.append(
        ["TOTAL — pallet 1", "", "", str(sum(c["qty"] for c in sub)),
         plain(round(sum(c["net"] for c in sub), 2)),
         plain(round(sum(c["gross"] for c in sub), 2))]
    )

    st = [
        *FACTORY_HEAD(),
        *doc_title("Pack-out sheet — pallet 1", f"{O.STYLE} · {O.PO_NO} · shipment 1 of 3"),
        two_col(
            [
                kv_block(
                    [
                        ("Buyer", O.BUYER["name"]),
                        ("PO", O.PO_NO),
                        ("Style", f"{O.STYLE} / {O.BUYER_ARTICLE}"),
                        ("Colour", "Charcoal Melange"),
                    ],
                    widths=(28, 58),
                )
            ],
            [
                kv_block(
                    [
                        ("Pack", f"{O.PCS_PER_CARTON} pcs solid colour solid size"),
                        ("Carton", "600 × 400 × 350 mm, 5-ply"),
                        ("Pallet", f"1 of {O.SHIP1_CARTONS // O.PACKOUT_CARTONS}"),
                        ("Packed on", "2027-01-16"),
                    ],
                    widths=(28, 58),
                )
            ],
        ),
        Spacer(1, 8),
        grid(rows, [30, 42, 18, 20, 24, 26], align_right=[3, 4, 5], font_size=7.4, zebra=True),
        Spacer(1, 8),
        P(
            f"This sheet covers cartons NKA-0001 to NKA-{O.PACKOUT_CARTONS:04d} only. The full "
            f"{O.SHIP1_CARTONS:,}-carton list for shipment 1 is in "
            "23-packing-list-shipment-1.xlsx and goes to the bank with the documents.",
            "small",
        ),
        Spacer(1, 8),
        signature_row(["Packing in-charge", "Finishing QC", "Shipping desk"]),
    ]
    build_pdf(out / "22-packout-sheet-pallet-1.pdf", st, "Pack-out sheet pallet 1")

    write_text(
        out / "22-packout-sheet-pallet-1.paste.txt",
        f"""{O.FACTORY['name']}
PACK-OUT SHEET — PALLET 1
{O.STYLE} · {O.PO_NO} · shipment 1 of 3

Buyer: {O.BUYER['name']}
Colour: Charcoal Melange
Pack: {O.PCS_PER_CARTON} pcs solid colour solid size
Packed on: 2027-01-16

Carton no   Colour             Size   Qty   Net kg   Gross kg
"""
        + "\n".join(
            f"{c['no']:<12}{c['color']:<19}{c['size']:<7}{c['qty']:<6}"
            f"{plain(c['net']):<9}{plain(c['gross'])}"
            for c in sub
        )
        + f"""
TOTAL - pallet 1                      {sum(c['qty'] for c in sub)}   """
        f"""{plain(round(sum(c['net'] for c in sub), 2))}   """
        f"""{plain(round(sum(c['gross'] for c in sub), 2))}
""",
    )
    write_json(
        out / "22-packout-sheet-pallet-1.expected.json",
        {
            "_intakeKind": "packing_list",
            "_door": "/shipment → Pack → the drop zone",
            "reference": f"{O.PO_NO} shipment 1 pallet 1",
            "_cartonCount": len(sub),
            "cartons": [
                {
                    "cartonNo": c["no"],
                    "contents": [{"color": c["color"], "size": c["size"], "qty": c["qty"]}],
                    "grossKg": plain(c["gross"]),
                    "netKg": plain(c["net"]),
                }
                for c in sub
            ],
            "_notes": [
                f"{len(sub)} cartons, each with exactly ONE content line — solid colour, solid "
                "size. A carton that comes back with an assortment has invented one.",
                "The TOTAL row is not a carton. 41 cartons, or a carton numbered 'TOTAL', is "
                "the failure this document is shaped to catch.",
                "grossKg and netKg are decimal STRINGS matching ^\\d{1,8}(\\.\\d{1,3})?$ — "
                "'15.50' and '14.52'. A gross of 15.5 is tolerable; 15,50 or '15.50 kg' is not.",
                "Every carton is 24 pcs. The qty must never be read off the carton NUMBER.",
            ],
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 23 · the whole shipment's carton list (xlsx — the bank's copy)
# ─────────────────────────────────────────────────────────────────────────────


def packing_list_xlsx(out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cartons"
    ws["A1"] = f"{O.FACTORY['name']} — packing list"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"{O.BUYER['name']} · PO {O.PO_NO} · style {O.STYLE} · shipment 1 of 3 · "
        f"invoice {O.INVOICE_NO} · B/L {O.BL_NO}"
    )
    ws["A2"].font = Font(size=9, italic=True)
    ws["A3"] = (
        f"{O.SHIP1_QTY:,} pcs in {O.SHIP1_CARTONS:,} cartons · {O.POL} to {O.POD} · "
        f"vessel {O.VESSEL}"
    )
    ws["A3"].font = Font(size=9, italic=True)

    _head(ws, 5, ["Carton no", "Colour", "Size", "Qty (pcs)", "Net kg", "Gross kg",
                  "Dimensions (cm)", "CBM"])
    r = 6
    for c in ALL_CARTONS:
        vals = [c["no"], c["color"], c["size"], c["qty"], c["net"], c["gross"],
                "60 × 40 × 35", 0.084]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = BORDER
            if j in (5, 6):
                cell.number_format = "0.00"
            if j == 8:
                cell.number_format = "0.000"
        r += 1
    totals = [
        "TOTAL", "", "", sum(c["qty"] for c in ALL_CARTONS),
        round(sum(c["net"] for c in ALL_CARTONS), 2),
        round(sum(c["gross"] for c in ALL_CARTONS), 2),
        f"{len(ALL_CARTONS):,} cartons", round(0.084 * len(ALL_CARTONS), 3),
    ]
    for j, v in enumerate(totals, start=1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.font, cell.fill, cell.border = Font(bold=True), TOT_FILL, BORDER
        if j in (5, 6):
            cell.number_format = "0.00"
        if j == 8:
            cell.number_format = "0.000"

    # a size summary, the way a real packing list carries one
    r += 3
    ws.cell(row=r, column=1, value="Summary by size").font = Font(bold=True, size=11)
    r += 1
    _head(ws, r, ["Size", "Cartons", "Pcs", "Net kg", "Gross kg"])
    for i, s in enumerate(O.SIZES):
        rows_s = [c for c in ALL_CARTONS if c["size"] == s]
        vals = [s, len(rows_s), sum(c["qty"] for c in rows_s),
                round(sum(c["net"] for c in rows_s), 2),
                round(sum(c["gross"] for c in rows_s), 2)]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r + 1 + i, column=j, value=v)
            cell.border = BORDER
            if j in (4, 5):
                cell.number_format = "0.00"

    _style_sheet(ws, {"A": 13, "B": 20, "C": 8, "D": 11, "E": 10, "F": 11, "G": 18, "H": 9})
    ws.freeze_panes = "A6"
    p = out / "23-packing-list-shipment-1.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)


# ─────────────────────────────────────────────────────────────────────────────
# 24 · the commercial invoice (no door — the bank set)
# ─────────────────────────────────────────────────────────────────────────────


def commercial_invoice(out: Path) -> None:
    by_size = []
    for s in O.SIZES:
        rows_s = [c for c in ALL_CARTONS if c["size"] == s]
        qty = sum(c["qty"] for c in rows_s)
        if qty:
            by_size.append([s, f"{len(rows_s):,}", f"{qty:,}", money(O.UNIT_PRICE),
                            money(round(qty * O.UNIT_PRICE, 2))])
    rows = [["Size", "Cartons", "Quantity (pcs)", f"Unit price {O.CURRENCY}",
             f"Amount {O.CURRENCY}"]] + by_size
    rows.append(["TOTAL", f"{O.SHIP1_CARTONS:,}", f"{O.SHIP1_QTY:,}", "",
                 money(O.SHIP1_VALUE)])

    st = [
        *FACTORY_HEAD(),
        *doc_title("Commercial invoice", f"{O.INVOICE_NO} · {O.INVOICE_DATE}"),
        two_col(
            [
                P("<b>Consignee / buyer</b>", "small"),
                P(O.BUYER["name"], "b"),
                P(O.BUYER["addr"][0], "p"),
                P(O.BUYER["addr"][1], "p"),
                P(f"VAT SE556812349401", "small"),
                Spacer(1, 5),
                P("<b>Notify party</b>", "small"),
                P("Same as consignee", "p"),
            ],
            [
                kv_block(
                    [
                        ("Invoice no", O.INVOICE_NO),
                        ("Invoice date", O.INVOICE_DATE),
                        ("Buyer PO", O.PO_NO),
                        ("L/C no", f"{O.LC_NO} dt. {O.LC_ISSUE}"),
                        ("Issuing bank", O.LC_ISSUING_BANK),
                        ("EXP no", O.EXP_NO),
                        ("B/L no", f"{O.BL_NO} dt. {O.BL_DATE}"),
                        ("Vessel / voyage", O.VESSEL),
                        ("Port of loading", O.POL),
                        ("Port of discharge", O.POD),
                        ("Terms", O.PRICE_TERM),
                    ],
                    widths=(30, 56),
                )
            ],
        ),
        Spacer(1, 8),
        P(f"Description of goods: {O.STYLE_LONG}. Style {O.STYLE}, buyer article "
          f"{O.BUYER_ARTICLE}, colour Charcoal Melange. Country of origin: Bangladesh. "
          f"HS code 6110.20.00.", "p"),
        Spacer(1, 6),
        grid(rows, [24, 26, 38, 42, 46], align_right=[1, 2, 3, 4]),
        Spacer(1, 8),
        two_col(
            [
                P("Shipment", "h2"),
                kv_block(
                    [
                        ("Cartons", f"{O.SHIP1_CARTONS:,}"),
                        ("Net weight", f"{sum(c['net'] for c in ALL_CARTONS):,.2f} kg"),
                        ("Gross weight", f"{sum(c['gross'] for c in ALL_CARTONS):,.2f} kg"),
                        ("Volume", f"{0.084 * len(ALL_CARTONS):,.3f} CBM"),
                        ("Marks", f"{O.BUYER_ARTICLE} / {O.PO_NO} / MADE IN BANGLADESH"),
                    ],
                    widths=(28, 58),
                ),
            ],
            [
                P("Declaration", "h2"),
                P(
                    "We certify that this invoice is true and correct, that the goods are of "
                    "Bangladesh origin, and that they have been manufactured in compliance "
                    f"with the terms of L/C {O.LC_NO} and the buyer's Code of Conduct.<br/><br/>"
                    "Payment: 120 days from bill of lading date, to our account with "
                    f"{O.LC_ADVISING_BANK}, {O.LC_ADVISING_BRANCH}, "
                    f"SWIFT {O.LC_ADVISING_SWIFT}, A/C 0021-4471-9930.",
                    "p",
                ),
            ],
        ),
        Spacer(1, 8),
        signature_row(["For " + O.FACTORY["name"] + "\nAuthorised signatory"]),
    ]
    build_pdf(out / "24-commercial-invoice-TT-INV-2815-1.pdf", st, f"Invoice {O.INVOICE_NO}")


# ─────────────────────────────────────────────────────────────────────────────
# 25 · the EXP form — a bank-stamped scan, the gate before any bank document
# ─────────────────────────────────────────────────────────────────────────────


def exp_form(out: Path) -> None:
    sh = Sheet(seed="exp-form")
    W, m = sh.w, 150

    sh.text((W // 2, 110), "BANGLADESH BANK", F_SANS_B, 44, anchor="ma")
    sh.bn((W // 2, 172), "রপ্তানি ফরম (ইএক্সপি)", size=40, bold=True, anchor="ma")
    sh.text((W // 2, 236), "EXPORT FORM  —  EXP", F_SANS_B, 36, anchor="ma")
    sh.text((W // 2, 288), "To be completed in quadruplicate by the exporter and certified "
            "by the authorised dealer", F_SANS, 22, fill=(80, 80, 80), anchor="ma")
    sh.hrule(330, m, W - m, w=4)

    def fld(x, y, en, bn, val, w, size=30):
        sh.text((x, y), en, F_SANS, 24, fill=(88, 88, 88))
        if bn:
            sh.bn((x + sh.width(en + "  ", F_SANS, 24), y - 2), bn, size=23, fill=(88, 88, 88))
        sh.text((x + 8, y + 38), val, F_SANS_B, size)
        sh.line((x, y + 84), (x + w, y + 84), fill=(150, 150, 150), w=2)

    y = 380
    COL = 1000
    RX = m + 1180
    pairs_l = [
        ("EXP No.", "ইএক্সপি নং", O.EXP_NO),
        ("Date of issue", "ইস্যুর তারিখ", "2027-01-14"),
        ("Authorised dealer", "অনুমোদিত ডিলার", O.LC_ADVISING_BANK),
        ("AD code", "", "010-2244-1188"),
        ("Exporter", "রপ্তানিকারক", O.FACTORY["name"]),
        ("ERC / BIN", "", "RC-119847 / 004471003-0201"),
    ]
    pairs_r = [
        ("Consignee", "প্রাপক", O.BUYER["name"]),
        ("Country of destination", "গন্তব্য দেশ", "Sweden"),
        ("Port of shipment", "শিপমেন্ট বন্দর", O.POL),
        ("L/C no. and date", "ঋণপত্র", f"{O.LC_NO} / {O.LC_ISSUE}"),
        ("Invoice no. and date", "চালান", f"{O.INVOICE_NO} / {O.INVOICE_DATE}"),
        ("Bill of lading", "বিল অব লেডিং", f"{O.BL_NO} / {O.BL_DATE}"),
    ]
    for i, (en, bn, v) in enumerate(pairs_l):
        fld(m, y + i * 110, en, bn, v, COL, 30 if len(v) < 26 else 26)
    for i, (en, bn, v) in enumerate(pairs_r):
        fld(RX, y + i * 110, en, bn, v, W - m - RX, 30 if len(v) < 26 else 26)
    y += 6 * 110 + 30

    rows = [
        ["Commodity", "HS code", "Quantity", "Unit value", "Total value"],
        [f"{O.STYLE_DESC}\n(style {O.STYLE})", "6110.20.00", f"{O.SHIP1_QTY:,} pcs",
         f"USD {money(O.UNIT_PRICE)}", f"USD {money(O.SHIP1_VALUE)}"],
    ]
    x = m
    cw = [820, 320, 400, 380, 460]
    sh.d.rectangle([m, y, m + sum(cw), y + 76], fill=(228, 228, 222))
    for j, lab in enumerate(rows[0]):
        sh.d.rectangle([x, y, x + cw[j], y + 76], outline=(90, 90, 90), width=2)
        sh.text((x + cw[j] // 2, y + 38), lab, F_SANS_B, 26, anchor="mm")
        x += cw[j]
    y += 76
    x = m
    for j, cell in enumerate(rows[1]):
        sh.d.rectangle([x, y, x + cw[j], y + 120], outline=(90, 90, 90), width=2)
        for k, ln in enumerate(cell.split("\n")):
            sh.text((x + 14, y + 34 + k * 40), ln, F_SANS_B, 27)
        x += cw[j]
    y += 170

    sh.text((m, y), "Terms of sale", F_SANS, 24, fill=(88, 88, 88))
    sh.text((m + 8, y + 38), f"{O.PRICE_TERM} — proceeds realisable within 120 days of "
            "shipment", F_SANS_B, 27)
    sh.line((m, y + 84), (W - m, y + 84), fill=(150, 150, 150), w=2)
    y += 130

    sh.box((m, y), (W - 2 * m, 210), outline=(110, 110, 110), w=3)
    sh.text((m + 24, y + 22), "DECLARATION BY THE EXPORTER", F_SANS_B, 26)
    for i, ln in enumerate(
        [
            "I/We declare that the particulars given above are true, that the export proceeds will be",
            "repatriated to Bangladesh through the authorised dealer named above within the period",
            "prescribed by Bangladesh Bank, and that no part of the value has been or will be received",
            "otherwise than through banking channels.",
        ]
    ):
        sh.text((m + 24, y + 66 + i * 36), ln, F_SANS, 23, fill=(58, 58, 58))
    y += 260

    sh.sign((m + 60, y + 30), "exp-exporter", scale=0.85)
    sh.line((m, y + 118), (m + 640, y + 118), fill=(90, 90, 90), w=2)
    sh.text((m, y + 134), "Signature and seal of the exporter", F_SANS, 23, fill=(70, 70, 70))
    sh.sign((W - m - 600, y + 30), "exp-ad", scale=0.85, color=(20, 30, 80))
    sh.line((W - m - 700, y + 118), (W - m, y + 118), fill=(90, 90, 90), w=2)
    sh.text((W - m - 700, y + 134), "Certified — authorised dealer", F_SANS, 23, fill=(70, 70, 70))
    sh.text((W - m - 700, y + 168), f"{O.LC_ADVISING_BANK}, AD 010-2244-1188", F_SANS, 21,
            fill=(70, 70, 70))

    sh.stamp((m + 760, y - 60), ["KARNAPHULI", "AD 010-2244-1188", "2027-01-14"], r=168, rot=8)
    sh.text((m, sh.h - 130), "TEST FIXTURE — generated for FabricXAI platform testing. "
            "Not a real EXP form and carries no legal meaning.", F_SANS, 22, fill=(140, 140, 140))

    save_scan_pdf(sh, out / "25-exp-form-certified.pdf", "exp-scan", grain=12, dark=0.96)
    write_text(
        out / "25-exp-form-certified.paste.txt",
        f"""BANGLADESH BANK — EXPORT FORM (EXP)

EXP No.: {O.EXP_NO}
Date of issue: 2027-01-14
Authorised dealer: {O.LC_ADVISING_BANK}
AD code: 010-2244-1188
Exporter: {O.FACTORY['name']}
ERC / BIN: RC-119847 / 004471003-0201
Consignee: {O.BUYER['name']}
Country of destination: Sweden
Port of shipment: {O.POL}
L/C no. and date: {O.LC_NO} / {O.LC_ISSUE}
Invoice no. and date: {O.INVOICE_NO} / {O.INVOICE_DATE}
Bill of lading: {O.BL_NO} / {O.BL_DATE}

Commodity: {O.STYLE_DESC} (style {O.STYLE})
HS code: 6110.20.00
Quantity: {O.SHIP1_QTY:,} pcs
Unit value: USD {money(O.UNIT_PRICE)}
Total value: USD {money(O.SHIP1_VALUE)}
Terms of sale: {O.PRICE_TERM} - proceeds realisable within 120 days of shipment
""",
    )


def build(out: Path) -> None:
    packout_sheet(out)
    packing_list_xlsx(out)
    commercial_invoice(out)
    exp_form(out)
