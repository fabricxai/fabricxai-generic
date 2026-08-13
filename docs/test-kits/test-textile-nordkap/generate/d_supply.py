"""Procurement and store: the proformas, the roll list, the two challans at the gate."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.platypus import Spacer

import _order as O
from _lib import (
    F_SANS,
    F_SANS_B,
    P,
    Sheet,
    build_pdf,
    doc_title,
    grid,
    kv_block,
    letterhead,
    money,
    save_photo,
    signature_row,
    two_col,
    write_json,
    write_text,
)
from d_merch import BORDER, HEAD_FILL, TOT_FILL, _head, _style_sheet


# ─────────────────────────────────────────────────────────────────────────────
# 10 · the mill's proforma (door: supplier_proforma → supplier_quotes)
# ─────────────────────────────────────────────────────────────────────────────


def fabric_proforma(out: Path) -> None:
    st = [
        *letterhead(
            O.MILL["name"], O.MILL["addr"], O.MILL["contact"] + ["Est. 1998 · ISO 9001:2015"],
            accent="#7a1f1f",
        ),
        *doc_title("Proforma invoice", f"{O.PI_NO} · {O.PI_DATE}"),
        two_col(
            [
                P("<b>To (buyer)</b>", "small"),
                P(O.FACTORY["name"], "b"),
                P(O.FACTORY["addr"][0], "p"),
                P(O.FACTORY["addr"][1], "p"),
                P(f"Attn: Procurement · ref your enquiry TT-ENQ-2815-F", "small"),
            ],
            [
                kv_block(
                    [
                        ("PI number", O.PI_NO),
                        ("PI date", O.PI_DATE),
                        ("Valid until", O.PI_VALID),
                        ("Price term", "CFR Chattogram"),
                        ("Currency", "USD"),
                        ("For end style", f"{O.STYLE} / {O.PO_NO}"),
                    ],
                    widths=(28, 58),
                )
            ],
        ),
        Spacer(1, 9),
        grid(
            [
                ["Article", "Description", "Qty", "Unit", "Unit price", "Amount"],
                [
                    "FAB-FLC-280",
                    P("Brushed back fleece 280 g/m² ±5%, 80% cotton / 20% polyester, "
                      "tubular 185 cm, reactive dyed, enzyme washed, brushed one side. "
                      "Shades: Charcoal Melange, Deep Navy, Off White.", "small"),
                    f"{O.FLEECE_KG:,}",
                    "kg",
                    f"{money(O.FLEECE_RATE)}",
                    f"{money(O.FLEECE_VALUE)}",
                ],
                ["", P("<b>Total CFR Chattogram</b>", "b"), "", "", "", f"{money(O.FLEECE_VALUE)}"],
            ],
            [24, 78, 20, 14, 22, 24],
            align_right=[2, 4, 5],
        ),
        Spacer(1, 9),
        two_col(
            [
                P("Terms", "h2"),
                kv_block(
                    [
                        ("Payment", "back-to-back L/C at 120 days from B/L date"),
                        ("Lead time", f"{O.PI_LEAD_DAYS} days from receipt of workable L/C"),
                        ("Minimum order", f"{O.PI_MOQ:,} kg per shade"),
                        ("Ocean freight", f"USD {money(O.PI_FREIGHT)} — included in the CFR price"),
                        ("Shipment", "part shipment allowed, 3 lots by shade"),
                        ("Tolerance", "±3% on quantity and value"),
                        ("Port of loading", "Ningbo, China"),
                    ],
                    widths=(26, 60),
                ),
            ],
            [
                P("Quality and inspection", "h2"),
                P(
                    "GSM tested per ASTM D3776. Shrinkage max 5% length / 4% width after "
                    "3 washes to AATCC 135. Colour fastness to washing min 4 (ISO 105-C06), "
                    "rubbing dry 4 / wet 3.<br/><br/>"
                    "Fabric supplied roll-wise with a packing list stating roll number, "
                    "weight, lot and shade group. Buyer's 4-point inspection at 10% of rolls "
                    "is accepted; rolls above 20 points per 100 sq yd may be returned at our "
                    "cost within 30 days of arrival.<br/><br/>"
                    "OEKO-TEX Standard 100 Class II certificate issued per shipment.",
                    "p",
                ),
            ],
        ),
        Spacer(1, 8),
        P(
            "Bank details for L/C advising: Bank of Ningbo, Shaoxing Keqiao Sub-branch, "
            "SWIFT BKNBCN2N, A/C 7702 8841 9930 4471, beneficiary as named above.",
            "small",
        ),
        Spacer(1, 8),
        signature_row(["For " + O.MILL["name"] + "\nExport Sales"]),
    ]
    build_pdf(out / "10-fabric-proforma-HL-PI-26-0914.pdf", st, f"Proforma {O.PI_NO}")

    write_text(
        out / "10-fabric-proforma-HL-PI-26-0914.paste.txt",
        f"""{O.MILL['name']}
{O.MILL['addr'][0]}
{O.MILL['addr'][1]}

PROFORMA INVOICE — {O.PI_NO} · {O.PI_DATE}

To (buyer): {O.FACTORY['name']}, {O.FACTORY['addr'][0]}, {O.FACTORY['addr'][1]}
Attn: Procurement · ref your enquiry TT-ENQ-2815-F
PI number: {O.PI_NO}
PI date: {O.PI_DATE}
Valid until: {O.PI_VALID}
Price term: CFR Chattogram
Currency: USD
For end style: {O.STYLE} / {O.PO_NO}

Article       Description                                                     Qty      Unit  Unit price  Amount
FAB-FLC-280   Brushed back fleece 280 g/m2 +/-5%, 80% cotton / 20% polyester,
              tubular 185 cm, reactive dyed, enzyme washed, brushed one side.
              Shades: Charcoal Melange, Deep Navy, Off White.            {O.FLEECE_KG:>9,}   kg      {money(O.FLEECE_RATE)}   {money(O.FLEECE_VALUE)}
Total CFR Chattogram                                                                                    {money(O.FLEECE_VALUE)}

TERMS
Payment: back-to-back L/C at 120 days from B/L date
Lead time: {O.PI_LEAD_DAYS} days from receipt of workable L/C
Minimum order: {O.PI_MOQ:,} kg per shade
Ocean freight: USD {money(O.PI_FREIGHT)} - included in the CFR price
Shipment: part shipment allowed, 3 lots by shade
Tolerance: +/-3% on quantity and value
Port of loading: Ningbo, China
""",
    )
    write_json(
        out / "10-fabric-proforma-HL-PI-26-0914.expected.json",
        {
            "_intakeKind": "supplier_proforma",
            "_door": "/procurement → New quote → 'read a proforma'",
            "reference": O.PI_NO,
            "quotedOn": O.PI_DATE,
            "validUntil": O.PI_VALID,
            "currency": "USD",
            "priceTerm": "CFR Chattogram",
            "lines": [
                {
                    "itemCode": "FAB-FLC-280",
                    "itemName": "Brushed back fleece 280 g/m² ±5%, 80% cotton / 20% "
                                "polyester, tubular 185 cm, reactive dyed, enzyme washed, "
                                "brushed one side",
                    "qty": str(O.FLEECE_KG),
                    "unit": "kg",
                    "unitPrice": f"{O.FLEECE_RATE:.2f}",
                    "leadTimeDays": O.PI_LEAD_DAYS,
                    "moq": str(O.PI_MOQ),
                    "freight": f"{O.PI_FREIGHT:.2f}",
                }
            ],
            "_notes": [
                "leadTimeDays is stated in prose — '35 days from receipt of workable L/C'. "
                "The schema preprocesses a string to its digits, so 35 is right; '35 days' "
                "surviving as a string means the preprocess did not run.",
                "dutyPct is NOT on this paper. It must come back absent — a 0 here is an "
                "invention, and imported fabric under bond is duty-free for a reason that "
                "has nothing to do with a rate the mill quoted.",
                "freight is stated as included in CFR. 1850.00 is the stated figure; do not "
                "expect it added to the line total.",
            ],
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 11 · the trims quotation, in taka (door: supplier_proforma)
# ─────────────────────────────────────────────────────────────────────────────


def trims_quotation(out: Path) -> None:
    rows = [["Item ref", "Description", "Qty", "Unit", "Rate BDT", "Amount BDT", "MOQ", "Lead"]]
    total = 0.0
    for t in O.TRIM_QUOTE:
        amt = round(t["qty"] * t["rate"], 2)
        total += amt
        rows.append(
            [
                t["ref"], P(t["name"], "small"), f"{t['qty']:,}", t["unit"],
                money(t["rate"]), money(amt), f"{t['moq']:,}", f"{t['lead']} d",
            ]
        )
    rows.append(["", P("<b>Total, EXW Dhaka</b>", "b"), "", "", "", money(total), "", ""])

    st = [
        *letterhead(
            O.TRIMS["name"], O.TRIMS["addr"],
            O.TRIMS["contact"] + ["BIN 001882447-0101 · Accessories & packaging"],
            accent="#1d3f7a",
        ),
        *doc_title("Quotation", f"{O.TQ_NO} · {O.TQ_DATE}"),
        two_col(
            [
                P("<b>To</b>", "small"),
                P(O.FACTORY["name"], "b"),
                P("Attn: Procurement", "p"),
                P(f"Ref: your RFQ for style {O.STYLE}, {O.PO_NO}", "small"),
            ],
            [
                kv_block(
                    [
                        ("Quotation no", O.TQ_NO),
                        ("Date", O.TQ_DATE),
                        ("Valid until", O.TQ_VALID),
                        ("Price term", "EXW Dhaka"),
                        ("Currency", "BDT"),
                    ],
                    widths=(28, 58),
                )
            ],
        ),
        Spacer(1, 9),
        grid(rows, [24, 56, 18, 12, 18, 22, 18, 12], align_right=[2, 4, 5, 6], font_size=7.6),
        Spacer(1, 9),
        two_col(
            [
                P("Terms", "h2"),
                kv_block(
                    [
                        ("Payment", "30 days from delivery against challan"),
                        ("Delivery", "to your Ashulia gate, our transport"),
                        ("Part delivery", "allowed, minimum 25% per drop"),
                        ("Zipper origin", "YKK Bangladesh, Ishwardi EPZ"),
                        ("Sampling", "3 sets free, thereafter at cost"),
                    ],
                    widths=(28, 58),
                ),
            ],
            [
                P("Notes", "h2"),
                P(
                    "1. Zipper tape and puller dyed to your approved lab dip; add 7 days for "
                    "shade approval before bulk.<br/>"
                    "2. Woven main labels are size-specific — please confirm the size "
                    "breakdown before we cut the loom.<br/>"
                    "3. Prices hold for the quantities quoted; a drop below MOQ re-prices.<br/>"
                    "4. Rates exclude VAT, which will be charged at the rate applicable on "
                    "the date of the challan.",
                    "p",
                ),
            ],
        ),
        Spacer(1, 8),
        signature_row(["For " + O.TRIMS["name"] + "\nSales"]),
    ]
    build_pdf(out / "11-trims-quotation-DTH-Q-2026-337.pdf", st, f"Quotation {O.TQ_NO}")

    lines_txt = "\n".join(
        f"{t['ref']:<16}{t['name'][:52]:<54}{t['qty']:>9,}  {t['unit']:<5}"
        f"{money(t['rate']):>9}{money(round(t['qty'] * t['rate'], 2)):>14}"
        f"{t['moq']:>9,}  {t['lead']} d"
        for t in O.TRIM_QUOTE
    )
    write_text(
        out / "11-trims-quotation-DTH-Q-2026-337.paste.txt",
        f"""{O.TRIMS['name']}
{O.TRIMS['addr'][0]}

QUOTATION — {O.TQ_NO} · {O.TQ_DATE}

To: {O.FACTORY['name']}, Attn: Procurement
Ref: your RFQ for style {O.STYLE}, {O.PO_NO}
Quotation no: {O.TQ_NO}
Date: {O.TQ_DATE}
Valid until: {O.TQ_VALID}
Price term: EXW Dhaka
Currency: BDT

Item ref        Description                                              Qty   Unit  Rate BDT   Amount BDT     MOQ  Lead
{lines_txt}
Total, EXW Dhaka {'':<70}{money(total):>14}

TERMS
Payment: 30 days from delivery against challan
Delivery: to your Ashulia gate, our transport
Part delivery: allowed, minimum 25% per drop
Zipper origin: YKK Bangladesh, Ishwardi EPZ
Rates exclude VAT.
""",
    )
    write_json(
        out / "11-trims-quotation-DTH-Q-2026-337.expected.json",
        {
            "_intakeKind": "supplier_proforma",
            "_door": "/procurement → New quote → 'read a proforma'",
            "reference": O.TQ_NO,
            "quotedOn": O.TQ_DATE,
            "validUntil": O.TQ_VALID,
            "currency": "BDT",
            "priceTerm": "EXW Dhaka",
            "lines": [
                {
                    "itemCode": t["ref"],
                    "itemName": t["name"],
                    "qty": str(t["qty"]),
                    "unit": t["unit"],
                    "unitPrice": f"{t['rate']:.2f}",
                    "leadTimeDays": t["lead"],
                    "moq": str(t["moq"]),
                }
                for t in O.TRIM_QUOTE
            ],
            "_notes": [
                "Currency is BDT, not USD. The schema defaults to USD when it cannot read "
                "one — a quote that comes back in dollars at BDT rates prices a zipper at "
                "USD 34.50 and the costing studio will believe it.",
                "Five lines. Rates carry two decimals including 1.15 and 7.20; a rate that "
                "arrives as 1.2 or 7.2 has been through a float.",
                "freight and dutyPct are absent from this paper — both must come back absent.",
            ],
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 12 · the mill's roll list (xlsx — the store types from it, no AI door)
# ─────────────────────────────────────────────────────────────────────────────


def roll_list_xlsx(out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Roll list"
    ws["A1"] = f"{O.MILL['name']} — packing list / roll wise detail"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"Challan {O.GRN_CHALLAN} · lot {O.GRN_LOT} · FAB-FLC-280 brushed fleece 280 g/m² · "
        f"Charcoal Melange · {O.GRN_ROLLS} rolls · {O.GRN_KG:,.1f} kg · for {O.STYLE} / {O.PO_NO}"
    )
    ws["A2"].font = Font(size=9, italic=True)
    ws["A3"] = (
        f"Under back-to-back credit {O.BTB1_NO} · UD {O.UD_NO} · BONDED — this material may "
        "not be issued without a UD reference"
    )
    ws["A3"].font = Font(size=9, italic=True, color="8A3A12")

    _head(ws, 5, ["Roll no", "Net kg", "Lot", "Shade group", "Width (cm)", "Mill 4-point", "Remark"])
    r = 6
    fail_fill = PatternFill("solid", fgColor="FBE3E0")
    for roll, kg, shade in O.ROLLS:
        pts = {"R-F-17": 24, "R-F-44": 27, "R-F-58": 22}.get(roll, None)
        vals = [
            roll, kg, O.GRN_LOT, shade, 185,
            pts if pts is not None else "",
            "FAILED at mill — 4-point above 20 pts/100 sq yd" if pts else "",
        ]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            if j == 2:
                c.number_format = "0.0"
            if pts:
                c.fill = fail_fill
        r += 1
    for col, val in ((1, "Total"), (2, round(O.GRN_KG, 1))):
        c = ws.cell(row=r, column=col, value=val)
        c.font, c.fill, c.border = Font(bold=True), TOT_FILL, BORDER
        if col == 2:
            c.number_format = "0.0"

    r += 2
    ws.cell(row=r, column=1, value=(
        f"Shade grouping: {O.ROLLS[0][0]} to R-F-{O.GRN_SHADE_B_FROM - 1:02d} are shade group A; "
        f"R-F-{O.GRN_SHADE_B_FROM:02d} to {O.ROLLS[-1][0]} are shade group B. "
        "Do not mix groups within one garment."
    )).font = Font(italic=True, size=9)
    r += 1
    ws.cell(row=r, column=1, value=(
        f"Three rolls ({', '.join(O.GRN_FAILED)}) failed the mill's own 4-point check and are "
        "shipped for claim settlement. They are NOT to be issued to production."
    )).font = Font(italic=True, size=9, color="8A3A12")

    _style_sheet(ws, {"A": 12, "B": 10, "C": 12, "D": 13, "E": 12, "F": 13, "G": 48})
    ws.freeze_panes = "A6"
    p = out / "12-mill-packing-list-roll-wise.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)


# ─────────────────────────────────────────────────────────────────────────────
# 13 · the fabric challan at the gate — a photograph (door: delivery_challan)
# ─────────────────────────────────────────────────────────────────────────────


def _challan_sheet(
    seed: str,
    supplier: dict,
    challan_no: str,
    date: str,
    lines: list[tuple[str, str, str, str]],
    note_en: str,
    note_bn: str,
    vehicle: str,
    handwritten: dict,
) -> Sheet:
    sh = Sheet((2480, 1754), seed=seed)   # A4 landscape — how challan books print
    W = sh.w
    m = 110

    sh.text((m, 80), supplier["name"], F_SANS_B, 46)
    sh.text((m, 142), supplier["addr"][0], F_SANS, 26, fill=(70, 70, 70))
    if len(supplier["addr"]) > 1:
        sh.text((m, 180), supplier["addr"][1], F_SANS, 26, fill=(70, 70, 70))
    sh.text((W - m, 80), "DELIVERY CHALLAN", F_SANS_B, 44, anchor="ra")
    sh.bn((W - m, 136), "ডেলিভারি চালান", size=38, bold=True, anchor="ra")
    sh.text((W - m, 194), "Original — Consignee copy", F_SANS, 24, fill=(90, 90, 90), anchor="ra")
    sh.hrule(238, m, W - m, w=4)

    # header fields, the two-column way a challan book is printed
    y = 274
    sh.text((m, y), "Challan No.", F_SANS, 26, fill=(85, 85, 85))
    sh.bn((m + 190, y - 2), "চালান নং", size=26, fill=(85, 85, 85))
    sh.hand((m + 340, y - 6), challan_no, size=40, bold=True)
    sh.text((m + 1250, y), "Date", F_SANS, 26, fill=(85, 85, 85))
    sh.bn((m + 1340, y - 2), "তারিখ", size=26, fill=(85, 85, 85))
    sh.hand((m + 1470, y - 6), date, size=38)

    y += 66
    sh.text((m, y), "Consignee", F_SANS, 26, fill=(85, 85, 85))
    sh.bn((m + 190, y - 2), "প্রাপক", size=26, fill=(85, 85, 85))
    sh.text((m + 340, y - 2), O.FACTORY["name"] + ", " + O.FACTORY["addr"][0], F_SANS_B, 28)
    y += 60
    sh.text((m, y), "Vehicle No.", F_SANS, 26, fill=(85, 85, 85))
    sh.bn((m + 190, y - 2), "গাড়ি নং", size=26, fill=(85, 85, 85))
    sh.hand((m + 340, y - 6), vehicle, size=36)
    sh.text((m + 1250, y), "Order ref", F_SANS, 26, fill=(85, 85, 85))
    sh.bn((m + 1400, y - 2), "অর্ডার", size=26, fill=(85, 85, 85))
    sh.hand((m + 1530, y - 6), f"{O.STYLE} / {O.PO_NO}", size=34)

    # goods table
    y += 74
    rows = [["Sl\nক্রমিক", "Description of goods / পণ্যের বিবরণ", "Quantity\nপরিমাণ", "Unit\nএকক"]]
    head_y = y
    col_w = [110, 1500, 380, 270]
    # header drawn by hand so the Bangla shapes
    x = m
    sh.d.rectangle([m, y, m + sum(col_w), y + 96], fill=(228, 228, 222))
    heads = [
        ("Sl", "ক্রমিক"), ("Description of goods", "পণ্যের বিবরণ"),
        ("Quantity", "পরিমাণ"), ("Unit", "একক"),
    ]
    for j, (en, bn) in enumerate(heads):
        sh.text((x + col_w[j] // 2, y + 16), en, F_SANS_B, 26, anchor="ma")
        sh.bn((x + col_w[j] // 2, y + 52), bn, size=24, fill=(60, 60, 60), anchor="ma")
        sh.d.rectangle([x, y, x + col_w[j], y + 96], outline=(90, 90, 90), width=2)
        x += col_w[j]
    y += 96
    for i, (sl, desc, qty, unit) in enumerate(lines):
        x = m
        for j, cell in enumerate([sl, desc, qty, unit]):
            sh.d.rectangle([x, y, x + col_w[j], y + 86], outline=(90, 90, 90), width=2)
            if j == 1:
                sh.text((x + 16, y + 43), cell, F_SANS, 28, anchor="lm")
            elif j == 0:
                sh.text((x + col_w[j] // 2, y + 43), cell, F_SANS, 28, anchor="mm")
            else:
                sh.hand((x + 24, y + 43), cell, size=34, anchor="lm")
            x += col_w[j]
        y += 86
    # a couple of empty ruled lines, as every challan book has
    for _ in range(2):
        x = m
        for j in range(4):
            sh.d.rectangle([x, y, x + col_w[j], y + 86], outline=(90, 90, 90), width=2)
            x += col_w[j]
        y += 86

    y += 40
    sh.text((m, y), note_en, F_SANS, 26, fill=(55, 55, 55))
    sh.bn((m, y + 42), note_bn, size=26, fill=(55, 55, 55))

    # handwritten gate scribble — what the storekeeper adds on receipt
    if handwritten:
        sh.hand((m + 1700, y - 20), handwritten["text"], size=32, fill=(150, 30, 30))

    # signatures
    sy = sh.h - 250
    sh.sign((m + 40, sy), seed + "-supplier", scale=0.8)
    sh.line((m, sy + 92), (m + 560, sy + 92), fill=(90, 90, 90), w=2)
    sh.text((m, sy + 106), "Authorised signatory (supplier)", F_SANS, 24, fill=(70, 70, 70))
    sh.bn((m, sy + 142), "সরবরাহকারীর স্বাক্ষর", size=24, fill=(70, 70, 70))

    sh.line((W // 2 - 280, sy + 92), (W // 2 + 280, sy + 92), fill=(90, 90, 90), w=2)
    sh.text((W // 2 - 280, sy + 106), "Driver", F_SANS, 24, fill=(70, 70, 70))
    sh.bn((W // 2 - 280, sy + 142), "চালক", size=24, fill=(70, 70, 70))

    sh.sign((W - m - 500, sy), seed + "-store", scale=0.8, color=(150, 30, 30))
    sh.line((W - m - 560, sy + 92), (W - m, sy + 92), fill=(90, 90, 90), w=2)
    sh.text((W - m - 560, sy + 106), "Received by (store)", F_SANS, 24, fill=(70, 70, 70))
    sh.bn((W - m - 560, sy + 142), "গ্রহণকারীর স্বাক্ষর ও সীল", size=24, fill=(70, 70, 70))

    sh.stamp((W - 1150, sy - 190), [supplier["name"].split()[0].upper(), "DHAKA", date],
             r=130, rot=-9, color=(140, 40, 40))
    sh.text((m, sh.h - 56), "TEST FIXTURE — FabricXAI platform testing. Not a real challan.",
            F_SANS, 20, fill=(150, 150, 150))
    return sh


def fabric_challan(out: Path) -> None:
    sh = _challan_sheet(
        "challan-fabric",
        {"name": O.MILL["name"], "addr": [
            "c/o Meghna Freight & Logistics Ltd (C&F 118/2019)",
            "Bonded transfer from Chattogram Port — BE No. C-884712 dt. 2026-11-09",
        ]},
        O.GRN_CHALLAN,
        O.GRN_DATE,
        [
            ("1", "FAB-FLC-280  Brushed fleece 280 gsm — Charcoal Melange",
             f"{O.GRN_KG:,.1f}", "kg"),
            ("2", f"Rolls R-F-01 to R-F-{O.GRN_ROLLS:02d}, lot {O.GRN_LOT}",
             f"{O.GRN_ROLLS}", "rolls"),
        ],
        "BONDED GOODS — received against UD " + O.UD_NO + " and back-to-back L/C " + O.BTB1_NO
        + ". Roll-wise packing list attached.",
        "বন্ডেড পণ্য — ইউডি ও ব্যাক-টু-ব্যাক ঋণপত্রের বিপরীতে গৃহীত।",
        "DHAKA METRO-T-11-4471",
        {"text": "3 rolls damaged — see QC"},
    )
    save_photo(sh, out / "13-fabric-challan-ZJH-DC-8842.jpg", "photo-fabric-challan")

    write_text(
        out / "13-fabric-challan-ZJH-DC-8842.paste.txt",
        f"""{O.MILL['name']}
c/o Meghna Freight & Logistics Ltd (C&F 118/2019)
Bonded transfer from Chattogram Port — BE No. C-884712 dt. 2026-11-09

DELIVERY CHALLAN / ডেলিভারি চালান
Original — Consignee copy

Challan No.: {O.GRN_CHALLAN}
Date: {O.GRN_DATE}
Consignee: {O.FACTORY['name']}, {O.FACTORY['addr'][0]}
Vehicle No.: DHAKA METRO-T-11-4471
Order ref: {O.STYLE} / {O.PO_NO}

Sl  Description of goods                                             Quantity    Unit
1   FAB-FLC-280  Brushed fleece 280 gsm - Charcoal Melange           {O.GRN_KG:,.1f}     kg
2   Rolls R-F-01 to R-F-{O.GRN_ROLLS:02d}, lot {O.GRN_LOT}                          {O.GRN_ROLLS}          rolls

BONDED GOODS - received against UD {O.UD_NO} and back-to-back L/C {O.BTB1_NO}.
Roll-wise packing list attached.
""",
    )
    write_json(
        out / "13-fabric-challan-ZJH-DC-8842.expected.json",
        {
            "_intakeKind": "delivery_challan",
            "_door": "/store/receive → the drop zone (fills the GRN form; nothing is queued)",
            "challanNo": O.GRN_CHALLAN,
            "receivedAt": O.GRN_DATE,
            "supplierName": O.MILL["name"],
            "lines": [
                {
                    "itemCode": "FAB-FLC-280",
                    "itemName": "Brushed fleece 280 gsm — Charcoal Melange",
                    "qty": f"{O.GRN_KG:.1f}",
                    "unit": "kg",
                    "rolls": [],
                }
            ],
            "_notes": [
                "This is a PHOTOGRAPH, not a scan and not a clean PDF. Attach the .jpg with "
                "nothing pasted, and the confidence should visibly pay for the light and the "
                "skew. Then run it again with the .paste.txt and compare the two maps — that "
                "difference IS the feature.",
                "ONE line, not two. Row 2 of the challan restates row 1 as a roll count, the "
                "way challan books do; reading it as a second material creates a phantom "
                "'rolls' item in the store.",
                "The 60 individual rolls are NOT on this challan — they are on the mill's "
                "packing list (12-mill-packing-list-roll-wise.xlsx). rolls[] should come back "
                "empty here. Type them from the spreadsheet, or the receipt has no roll "
                "traceability and the shade groups never exist.",
                "The red handwriting ('3 rolls damaged — see QC') is the storekeeper's own "
                "note. It is not a field. If it lands in a quantity, that is a finding.",
            ],
        },
    )


def trims_challan(out: Path) -> None:
    sh = _challan_sheet(
        "challan-trims",
        O.TRIMS,
        O.TRIMS_CHALLAN,
        O.TRIMS_CHALLAN_DATE,
        [
            ("1", "TRM-ZIP-OE65  YKK #5 open-end zipper 65 cm — assorted shades",
             f"{O.ZIP_PCS:,}", "pcs"),
            ("2", "TRM-CORD-8  Flat drawcord 8 mm x 130 cm, tipped", f"{O.CORD_PCS:,}", "pcs"),
            ("3", "TRM-EYELET-8  Metal eyelet 8 mm antique silver", f"{O.EYELET_PCS:,}", "pcs"),
            ("4", "TRM-LBL-MAIN  Woven main label, size-specific", f"{O.LABEL_PCS:,}", "pcs"),
        ],
        "General (non-bonded) store. Against our quotation " + O.TQ_NO + ". VAT challan follows.",
        "সাধারণ স্টোর — বন্ডেড নয়। ভ্যাট চালান পরে পাঠানো হইবে।",
        "DHAKA METRO-N-19-8802",
        {},
    )
    save_photo(sh, out / "14-trims-challan-DTH-4512.jpg", "photo-trims-challan")

    write_text(
        out / "14-trims-challan-DTH-4512.paste.txt",
        f"""{O.TRIMS['name']}
{O.TRIMS['addr'][0]}

DELIVERY CHALLAN / ডেলিভারি চালান

Challan No.: {O.TRIMS_CHALLAN}
Date: {O.TRIMS_CHALLAN_DATE}
Consignee: {O.FACTORY['name']}, {O.FACTORY['addr'][0]}
Vehicle No.: DHAKA METRO-N-19-8802
Order ref: {O.STYLE} / {O.PO_NO}

Sl  Description of goods                                             Quantity    Unit
1   TRM-ZIP-OE65  YKK #5 open-end zipper 65 cm - assorted shades     {O.ZIP_PCS:,}      pcs
2   TRM-CORD-8  Flat drawcord 8 mm x 130 cm, tipped                  {O.CORD_PCS:,}      pcs
3   TRM-EYELET-8  Metal eyelet 8 mm antique silver                   {O.EYELET_PCS:,}      pcs
4   TRM-LBL-MAIN  Woven main label, size-specific                    {O.LABEL_PCS:,}      pcs

General (non-bonded) store. Against our quotation {O.TQ_NO}. VAT challan follows.
""",
    )
    write_json(
        out / "14-trims-challan-DTH-4512.expected.json",
        {
            "_intakeKind": "delivery_challan",
            "_door": "/store/receive → the drop zone",
            "challanNo": O.TRIMS_CHALLAN,
            "receivedAt": O.TRIMS_CHALLAN_DATE,
            "supplierName": O.TRIMS["name"],
            "lines": [
                {"itemCode": "TRM-ZIP-OE65",
                 "itemName": "YKK #5 open-end zipper 65 cm — assorted shades",
                 "qty": str(O.ZIP_PCS), "unit": "pcs", "rolls": []},
                {"itemCode": "TRM-CORD-8",
                 "itemName": "Flat drawcord 8 mm x 130 cm, tipped",
                 "qty": str(O.CORD_PCS), "unit": "pcs", "rolls": []},
                {"itemCode": "TRM-EYELET-8",
                 "itemName": "Metal eyelet 8 mm antique silver",
                 "qty": str(O.EYELET_PCS), "unit": "pcs", "rolls": []},
                {"itemCode": "TRM-LBL-MAIN",
                 "itemName": "Woven main label, size-specific",
                 "qty": str(O.LABEL_PCS), "unit": "pcs", "rolls": []},
            ],
            "_notes": [
                "Four lines, all in pcs, all general store — nothing here is bonded, so the "
                "receipt must NOT ask for a UD. If it does, the bonded test in the store has "
                "the wrong default.",
                "86,520 (eyelets) is the largest number on the page and is two per garment "
                "plus wastage. A reading of 8,652 or 865,200 is the digit-slip this document "
                "exists to catch.",
            ],
        },
    )


def build(out: Path) -> None:
    fabric_proforma(out)
    trims_quotation(out)
    roll_list_xlsx(out)
    fabric_challan(out)
    trims_challan(out)
