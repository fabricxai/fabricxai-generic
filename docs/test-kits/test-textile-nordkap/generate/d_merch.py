"""Merchandising: the enquiry, the PO, the grid, the tech pack, the amendment."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from reportlab.lib.units import mm
from reportlab.platypus import Image as RLImage, KeepTogether, PageBreak, Spacer

import _order as O
from _lib import (
    F_SANS,
    F_SANS_B,
    P,
    Sheet,
    build_pdf,
    doc_title,
    grid,
    kv_block,
    letterhead,
    money,
    signature_row,
    two_col,
    write_json,
    write_text,
)

BUYER_HEAD = lambda: letterhead(
    O.BUYER["name"], O.BUYER["addr"], O.BUYER["contact"], accent=O.BUYER["accent"]
)
FACTORY_HEAD = lambda: letterhead(
    O.FACTORY["name"], O.FACTORY["addr"], O.FACTORY["contact"], accent="#1f4d2b"
)


# ─────────────────────────────────────────────────────────────────────────────
# 01 · the enquiry (door: buyer_enquiry → rfqs)
# ─────────────────────────────────────────────────────────────────────────────


def enquiry(out: Path) -> None:
    st = [
        *BUYER_HEAD(),
        *doc_title("Request for quotation", f"{O.ENQ_NO} · {O.SEASON}"),
        two_col(
            [
                P("<b>To</b>", "small"),
                P(O.FACTORY["name"], "b"),
                P("Attn: Merchandising — knitwear desk", "p"),
            ],
            [
                kv_block(
                    [
                        ("Enquiry no", O.ENQ_NO),
                        ("Date", O.ENQ_DATE),
                        ("Quote required by", O.ENQ_DEADLINE),
                        ("Raised by", f"{O.BUYER['person']}, {O.BUYER['person_title']}"),
                    ],
                    widths=(30, 55),
                )
            ],
        ),
        Spacer(1, 8),
        P(
            "We are opening a new knitwear vendor for AW-27 core and would like your price "
            "for the programme below. This is a first enquiry — we have not placed with you "
            "before, so please quote as a new-vendor programme and include your nominated "
            "trims positions separately.",
            "p",
        ),
        Spacer(1, 8),
        grid(
            [
                ["Item", "Detail"],
                ["Product type", "ladies' knitted hooded sweatshirt, full zip"],
                ["Our article", O.BUYER_ARTICLE],
                ["Your style ref", f"to be advised — quote against {O.BUYER_ARTICLE}"],
                ["Description", P(O.STYLE_LONG)],
                ["Quantity", f"{O.QTY:,} pcs total, 3 colours"],
                ["Size range", "XS – XL, ratio 1 : 2 : 3 : 2 : 1"],
                ["Fabric", "brushed back fleece 280 g/m², 80% cotton / 20% polyester"],
                ["Target price", f"USD {money(O.ENQ_TARGET_PRICE)} per piece, {O.PRICE_TERM}"],
                ["Requested ship", "last week of January 2027, ex-factory"],
                ["Payment", f"irrevocable L/C at {O.DELIVERY_TERM_DAYS} days from B/L date"],
                ["Compliance", "our Code of Conduct audit must close before order placement"],
            ],
            [42, 138],
            font_size=8.5,
        ),
        Spacer(1, 8),
        P(
            "Please break your price into fabric, trims, CM and commercial so we can compare "
            "like for like, and state your earliest ex-factory for the full quantity. Costed "
            "tech pack follows on confirmation of interest.",
            "p",
        ),
        Spacer(1, 12),
        P("Kind regards,", "p"),
        Spacer(1, 12),
        P(f"<b>{O.BUYER['person']}</b>", "p"),
        P(f"{O.BUYER['person_title']} · {O.BUYER['name']}", "small"),
    ]
    build_pdf(out / "01-buyer-enquiry-NKA-ENQ-4471.pdf", st, "Buyer enquiry NKA-ENQ-4471")

    write_text(
        out / "01-buyer-enquiry-NKA-ENQ-4471.paste.txt",
        f"""{O.BUYER['name']}
{O.BUYER['addr'][0]}
{O.BUYER['addr'][1]}

REQUEST FOR QUOTATION — {O.ENQ_NO} · {O.SEASON}

To: {O.FACTORY['name']}, Attn: Merchandising — knitwear desk
Enquiry no: {O.ENQ_NO}
Date: {O.ENQ_DATE}
Quote required by: {O.ENQ_DEADLINE}
Raised by: {O.BUYER['person']}, {O.BUYER['person_title']}

We are opening a new knitwear vendor for AW-27 core and would like your price for the
programme below.

Product type: ladies' knitted hooded sweatshirt, full zip
Our article: {O.BUYER_ARTICLE}
Description: {O.STYLE_LONG}
Quantity: {O.QTY:,} pcs total, 3 colours
Size range: XS - XL, ratio 1 : 2 : 3 : 2 : 1
Fabric: brushed back fleece 280 g/m2, 80% cotton / 20% polyester
Target price: USD {money(O.ENQ_TARGET_PRICE)} per piece, {O.PRICE_TERM}
Requested ship: last week of January 2027, ex-factory
Payment: irrevocable L/C at {O.DELIVERY_TERM_DAYS} days from B/L date
Compliance: our Code of Conduct audit must close before order placement
""",
    )
    write_json(
        out / "01-buyer-enquiry-NKA-ENQ-4471.expected.json",
        {
            "_intakeKind": "buyer_enquiry",
            "_door": "/marbim/intake → 'A buyer enquiry'",
            "_context": {
                "buyerId": "picked from the Buyer dropdown — Nordkap Apparel AB (NKA). "
                "Never in the document. Create the buyer first (lead → convert)."
            },
            "title": "ladies' brushed-fleece full-zip hoodie — AW-27 core",
            "productType": "ladies' knitted hooded sweatshirt, full zip",
            "quantity": O.QTY,
            "unit": "pcs",
            "targetPrice": f"{O.ENQ_TARGET_PRICE:.2f}",
            "targetCurrency": "USD",
            "currency": "USD",
            "sizeRatio": {"XS": 1, "S": 2, "M": 3, "L": 2, "XL": 1},
            "deadline": O.ENQ_DEADLINE,
            "_notes": [
                "styleCode is NOT on this paper — the buyer quotes their own article "
                f"{O.BUYER_ARTICLE} and asks the factory to advise a style. An extraction "
                "that invents ST-2815 here is a finding: the style did not exist yet.",
                "requestedShipDate is stated in prose ('last week of January 2027'), not as "
                "a date. Either 2027-01-25..31 or absent is defensible; a confident single "
                "date at high confidence is not.",
            ],
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 02 · the purchase order (door: buyer_po → orders)
# ─────────────────────────────────────────────────────────────────────────────


def _grid_rows() -> list[list[str]]:
    rows = [["Colour", "Colour code"] + O.SIZES + ["Total"]]
    for c in O.COLORS:
        rows.append(
            [c, O.COLOR_CODES[c]]
            + [f"{O.BREAKDOWN[c][s]:,}" for s in O.SIZES]
            + [f"{O.color_total(c):,}"]
        )
    rows.append(
        ["Total", ""] + [f"{O.size_total(s):,}" for s in O.SIZES] + [f"{O.QTY:,}"]
    )
    return rows


def purchase_order(out: Path) -> None:
    st = [
        *BUYER_HEAD(),
        *doc_title("Purchase order", f"{O.PO_NO} · page 1 of 1"),
        two_col(
            [
                P("<b>Supplier</b>", "small"),
                P(O.FACTORY["name"], "b"),
                P(O.FACTORY["addr"][0], "p"),
                P(O.FACTORY["addr"][1], "p"),
                P("Vendor code NK-BD-0219", "small"),
            ],
            [
                kv_block(
                    [
                        ("PO number", O.PO_NO),
                        ("PO date", O.PO_DATE),
                        ("Against enquiry", O.ENQ_NO),
                        ("Season", O.SEASON),
                        ("Buyer article", O.BUYER_ARTICLE),
                        ("Currency", O.CURRENCY),
                    ],
                    widths=(30, 55),
                )
            ],
        ),
        Spacer(1, 9),
        grid(
            [
                ["Style", "Description", "Qty (pcs)", "Unit price", "Amount"],
                [
                    O.STYLE,
                    P(f"{O.STYLE_DESC}<br/>{O.STYLE_LONG}"),
                    f"{O.QTY:,}",
                    f"{money(O.UNIT_PRICE)}",
                    f"{money(O.ORDER_VALUE)}",
                ],
                ["", "TOTAL FOB CHATTOGRAM (USD)", f"{O.QTY:,}", "", f"{money(O.ORDER_VALUE)}"],
            ],
            [20, 96, 22, 20, 22],
            align_right=[2, 3, 4],
        ),
        Spacer(1, 9),
        P("Colour and size breakdown", "h2"),
        grid(_grid_rows(), [40, 24, 16, 16, 16, 16, 16, 22], align_right=[2, 3, 4, 5, 6, 7]),
        Spacer(1, 9),
        two_col(
            [
                P("Delivery and terms", "h2"),
                kv_block(
                    [
                        ("Price term", O.PRICE_TERM),
                        ("Ex-factory date", O.EX_FACTORY),
                        ("Port of loading", O.POL),
                        ("Port of discharge", O.POD),
                        ("Payment", f"irrevocable L/C at {O.DELIVERY_TERM_DAYS} days from B/L"),
                        ("Partial shipment", "allowed"),
                        ("Packing", f"{O.PCS_PER_CARTON} pcs solid colour solid size per carton"),
                    ],
                    widths=(28, 58),
                ),
            ],
            [
                P("Conditions", "h2"),
                P(
                    f"1. Tech pack {O.STYLE} {O.TECHPACK_REV} dated {O.TECHPACK_DATE} governs "
                    "construction, measurements and packing.<br/>"
                    "2. PP sample approval is required before bulk cutting.<br/>"
                    "3. All trims from nominated suppliers; substitutions in writing only.<br/>"
                    "4. Final inspection at AQL 2.5 major / 4.0 minor, level GII, by our "
                    "nominated inspector.<br/>"
                    "5. Code of Conduct audit findings must be closed to plan.<br/>"
                    "6. Late shipment: 3% of order value per commenced week, air freight at "
                    "supplier's cost beyond 2 weeks.",
                    "p",
                ),
            ],
            widths=(88, 92),
        ),
        Spacer(1, 6),
        signature_row(
            [
                f"{O.BUYER['person']} · {O.BUYER['person_title']}\n{O.BUYER['name']}",
                "Accepted for and on behalf of\n" + O.FACTORY["name"],
            ]
        ),
    ]
    build_pdf(out / "02-buyer-po-NKA-PO-70318.pdf", st, f"Purchase order {O.PO_NO}")

    grid_txt = "\n".join(
        "  ".join(
            [f"{c:<18}", f"{O.COLOR_CODES[c]:<4}"]
            + [f"{O.BREAKDOWN[c][s]:>7,}" for s in O.SIZES]
            + [f"{O.color_total(c):>8,}"]
        )
        for c in O.COLORS
    )
    write_text(
        out / "02-buyer-po-NKA-PO-70318.paste.txt",
        f"""{O.BUYER['name']}
{O.BUYER['addr'][0]}
{O.BUYER['addr'][1]}

PURCHASE ORDER — {O.PO_NO}

Supplier: {O.FACTORY['name']}, {O.FACTORY['addr'][0]}, {O.FACTORY['addr'][1]}
Vendor code: NK-BD-0219
PO number: {O.PO_NO}
PO date: {O.PO_DATE}
Against enquiry: {O.ENQ_NO}
Season: {O.SEASON}
Buyer article: {O.BUYER_ARTICLE}
Currency: {O.CURRENCY}

Style      Description                                              Qty (pcs)  Unit price   Amount
{O.STYLE}    {O.STYLE_DESC}                       {O.QTY:>9,}     {money(O.UNIT_PRICE)}   {money(O.ORDER_VALUE)}
           {O.STYLE_LONG}
TOTAL FOB CHATTOGRAM (USD)                                          {O.QTY:>9,}              {money(O.ORDER_VALUE)}

COLOUR AND SIZE BREAKDOWN
Colour              Code       XS        S        M        L       XL     Total
{grid_txt}
Total                      {O.size_total('XS'):>7,}  {O.size_total('S'):>7,}  {O.size_total('M'):>7,}  {O.size_total('L'):>7,}  {O.size_total('XL'):>7,}  {O.QTY:>8,}

DELIVERY AND TERMS
Price term: {O.PRICE_TERM}
Ex-factory date: {O.EX_FACTORY}
Port of loading: {O.POL}
Port of discharge: {O.POD}
Payment: irrevocable L/C at {O.DELIVERY_TERM_DAYS} days from B/L
Partial shipment: allowed
Packing: {O.PCS_PER_CARTON} pcs solid colour solid size per carton

CONDITIONS
1. Tech pack {O.STYLE} {O.TECHPACK_REV} dated {O.TECHPACK_DATE} governs construction,
   measurements and packing.
2. PP sample approval is required before bulk cutting.
3. All trims from nominated suppliers; substitutions in writing only.
4. Final inspection at AQL 2.5 major / 4.0 minor, level GII, by our nominated inspector.
5. Code of Conduct audit findings must be closed to plan.
6. Late shipment: 3% of order value per commenced week, air freight at supplier's cost
   beyond 2 weeks.
""",
    )
    write_json(
        out / "02-buyer-po-NKA-PO-70318.expected.json",
        {
            "_intakeKind": "buyer_po",
            "_door": "/marbim/intake → \"A buyer's purchase order\" · or /orders → New order",
            "_context": {
                "buyerId": "picked from the Buyer dropdown — Nordkap Apparel AB (NKA); "
                "never in the document"
            },
            "poNumbers": [O.PO_NO],
            "totalValue": f"{O.ORDER_VALUE:.2f}",
            "currency": "USD",
            "plannedExFactoryDate": O.EX_FACTORY,
            "styles": [
                {
                    "styleCode": O.STYLE,
                    "description": O.STYLE_DESC,
                    "contractedQty": O.QTY,
                    "unitPrice": f"{O.UNIT_PRICE:.2f}",
                    "currency": "USD",
                    "breakdown": [
                        {"color": c, "size": s, "qty": O.BREAKDOWN[c][s]}
                        for c in O.COLORS
                        for s in O.SIZES
                    ],
                }
            ],
            "_notes": [
                "ONE style with a 15-cell breakdown — not 15 styles, and not 3. The grid is "
                "the most repetitive thing on the page and the schema has exactly one home "
                "for it (styles[0].breakdown).",
                "The three colours do NOT share a size ratio: Off White runs 1200/2400/3600/"
                "2700/1500. An extraction that applies Charcoal's ratio to all three colours "
                "produces a 44,100-piece order and is a finding.",
                "Cells must sum to contractedQty = 42,000. Check the sum on the approve "
                "screen before approving.",
            ],
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 03 · the grid as the buyer's planner keeps it (xlsx — no AI door)
# ─────────────────────────────────────────────────────────────────────────────

THIN = Side(style="thin", color="9A9A9A")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="E8E8E4")
TOT_FILL = PatternFill("solid", fgColor="F4F1E4")


def _style_sheet(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _head(ws, row: int, values: list, fill=HEAD_FILL) -> None:
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(bold=True, size=9)
        c.fill = fill
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def breakdown_xlsx(out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order grid"
    ws["A1"] = f"{O.BUYER['name']} — order grid and delivery schedule"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"PO {O.PO_NO} · style {O.STYLE} · article {O.BUYER_ARTICLE} · "
        f"{O.QTY:,} pcs · ex-factory {O.EX_FACTORY}"
    )
    ws["A2"].font = Font(size=9, italic=True)

    _head(ws, 4, ["Colour", "Code", *O.SIZES, "Total", "Value USD"])
    r = 5
    for c in O.COLORS:
        ws.cell(row=r, column=1, value=c).border = BORDER
        ws.cell(row=r, column=2, value=O.COLOR_CODES[c]).border = BORDER
        for i, s in enumerate(O.SIZES):
            cell = ws.cell(row=r, column=3 + i, value=O.BREAKDOWN[c][s])
            cell.number_format = "#,##0"
            cell.border = BORDER
        t = ws.cell(row=r, column=8, value=O.color_total(c))
        t.number_format = "#,##0"
        t.font = Font(bold=True)
        t.border = BORDER
        v = ws.cell(row=r, column=9, value=round(O.color_total(c) * O.UNIT_PRICE, 2))
        v.number_format = "#,##0.00"
        v.border = BORDER
        r += 1
    ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
    for i, s in enumerate(O.SIZES):
        cell = ws.cell(row=r, column=3 + i, value=O.size_total(s))
        cell.font, cell.number_format, cell.border, cell.fill = (
            Font(bold=True), "#,##0", BORDER, TOT_FILL,
        )
    for col, val, fmt in ((8, O.QTY, "#,##0"), (9, O.ORDER_VALUE, "#,##0.00")):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font, cell.number_format, cell.border, cell.fill = (
            Font(bold=True), fmt, BORDER, TOT_FILL,
        )
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2).border = BORDER

    r += 3
    ws.cell(row=r, column=1, value="Delivery schedule").font = Font(bold=True, size=11)
    r += 1
    _head(ws, r, ["Shipment", "Ex-factory", "Colour", "Qty (pcs)", "Cartons", "Value USD", "Status"])
    plan = [
        ("1 of 3", "2027-01-28", "Charcoal Melange", O.SHIP1_QTY),
        ("2 of 3", "2027-02-08", "Deep Navy", 14400),
        ("3 of 3", "2027-02-18", "Off White + balance", O.QTY - O.SHIP1_QTY - 14400),
    ]
    for i, (name, date, col, qty) in enumerate(plan):
        rr = r + 1 + i
        vals = [
            name, date, col, qty, qty // O.PCS_PER_CARTON,
            round(qty * O.UNIT_PRICE, 2),
            "booked" if i == 0 else "planned",
        ]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=rr, column=j, value=v)
            cell.border = BORDER
            if j in (4, 5):
                cell.number_format = "#,##0"
            if j == 6:
                cell.number_format = "#,##0.00"
    r += len(plan) + 3
    ws.cell(
        row=r,
        column=1,
        value=(
            "NOTE — shipment 3 falls on 2027-02-18, which is AFTER the L/C latest shipment "
            f"date of {O.LC_LATEST_SHIPMENT}. The buyer is aware and an L/C amendment is to "
            "follow. Do not present documents for shipment 3 until the amendment is advised."
        ),
    ).font = Font(italic=True, color="8A3A12", size=9)

    _style_sheet(ws, {"A": 22, "B": 8, "C": 10, "D": 10, "E": 10, "F": 10, "G": 10, "H": 12, "I": 14})
    ws.freeze_panes = "A5"

    p = out / "03-order-grid-and-delivery-schedule.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)


# ─────────────────────────────────────────────────────────────────────────────
# 04 · the tech pack (door: tech_pack → boms) with a flat sketch
# ─────────────────────────────────────────────────────────────────────────────


def _flat_sketch(path: Path) -> Path:
    """Front and back flats, drawn — a tech pack without a sketch reads as a form."""
    W, H = 1700, 980
    sh = Sheet((W, H), bg=(255, 255, 255), seed="flat")
    d = sh.d
    ink = (34, 34, 38)
    thin = (120, 120, 126)

    def hoodie(ox: int, back: bool) -> None:
        # body
        body = [
            (ox + 120, 250), (ox + 120, 700), (ox + 150, 760), (ox + 450, 760),
            (ox + 480, 700), (ox + 480, 250),
        ]
        d.line(body + [body[0]], fill=ink, width=5)
        # shoulders + sleeves
        d.line([(ox + 120, 250), (ox + 40, 300), (ox + 10, 620), (ox + 95, 645), (ox + 130, 340)],
               fill=ink, width=5)
        d.line([(ox + 480, 250), (ox + 560, 300), (ox + 590, 620), (ox + 505, 645), (ox + 470, 340)],
               fill=ink, width=5)
        # cuffs
        d.rectangle([ox + 10, 620, ox + 95, 680], outline=ink, width=5)
        d.rectangle([ox + 505, 620, ox + 590, 680], outline=ink, width=5)
        for i in range(8):
            d.line([(ox + 14 + i * 10, 624), (ox + 14 + i * 10, 676)], fill=thin, width=2)
            d.line([(ox + 509 + i * 10, 624), (ox + 509 + i * 10, 676)], fill=thin, width=2)
        # bottom rib
        d.rectangle([ox + 130, 700, ox + 470, 760], outline=ink, width=5)
        for i in range(33):
            d.line([(ox + 134 + i * 10, 704), (ox + 134 + i * 10, 756)], fill=thin, width=2)
        # hood
        d.arc([ox + 150, 60, ox + 450, 330], start=180, end=360, fill=ink, width=5)
        d.line([(ox + 150, 195), (ox + 175, 265)], fill=ink, width=5)
        d.line([(ox + 450, 195), (ox + 425, 265)], fill=ink, width=5)
        d.arc([ox + 185, 120, ox + 415, 300], start=180, end=360, fill=thin, width=3)
        if back:
            # CB print box + seam
            d.rectangle([ox + 230, 330, ox + 370, 430], outline=(180, 60, 60), width=4)
            sh.text((ox + 300, 445), "CB PRINT 180×120", F_SANS, 22,
                    fill=(180, 60, 60), anchor="ma")
            d.line([(ox + 300, 250), (ox + 300, 700)], fill=thin, width=2)
        else:
            # zip
            d.line([(ox + 300, 215), (ox + 300, 700)], fill=ink, width=5)
            for y in range(225, 700, 16):
                d.line([(ox + 292, y), (ox + 308, y)], fill=thin, width=2)
            d.rectangle([ox + 289, 300, ox + 311, 340], outline=ink, width=4)
            # kangaroo pocket
            d.line([(ox + 165, 560), (ox + 205, 500), (ox + 395, 500), (ox + 435, 560),
                    (ox + 435, 700)], fill=ink, width=4)
            d.line([(ox + 165, 560), (ox + 165, 700)], fill=ink, width=4)
            # drawcord
            d.line([(ox + 268, 232), (ox + 250, 300)], fill=ink, width=4)
            d.line([(ox + 332, 232), (ox + 350, 300)], fill=ink, width=4)
            d.ellipse([ox + 262, 224, ox + 276, 238], outline=ink, width=3)
            d.ellipse([ox + 326, 224, ox + 340, 238], outline=ink, width=3)

    hoodie(90, back=False)
    hoodie(920, back=True)
    sh.text((390, 850), "FRONT", F_SANS_B, 30, anchor="ma")
    sh.text((1220, 850), "BACK", F_SANS_B, 30, anchor="ma")
    sh.text((850, 910), f"{O.STYLE} · {O.STYLE_DESC} · {O.TECHPACK_REV}", F_SANS, 24,
            fill=(90, 90, 96), anchor="ma")
    path.parent.mkdir(parents=True, exist_ok=True)
    sh.img.save(path)
    return path


def tech_pack(out: Path) -> None:
    sketch = _flat_sketch(out / "04-tech-pack-ST-2815.flat-sketch.png")

    bom_rows = [["Group", "Item ref", "Specification", "Cons/pc", "UoM", "Wastage %", "Page"]]
    for b in O.BOM:
        bom_rows.append(
            [b["group"], b["ref"], P(b["spec"], "small"), b["cons"], b["uom"], b["waste"],
             f"p.{b['page']}"]
        )

    st = [
        *BUYER_HEAD(),
        *doc_title("Technical package", f"{O.STYLE} · {O.TECHPACK_REV} · {O.TECHPACK_DATE}"),
        two_col(
            [
                kv_block(
                    [
                        ("Style", O.STYLE),
                        ("Buyer article", O.BUYER_ARTICLE),
                        ("Description", O.STYLE_DESC),
                        ("Season", O.SEASON),
                    ],
                    widths=(26, 60),
                )
            ],
            [
                kv_block(
                    [
                        ("Revision", O.TECHPACK_REV),
                        ("Issued", O.TECHPACK_DATE),
                        ("Supersedes", "Rev 1 dated 2026-08-29"),
                        ("Against PO", O.PO_NO),
                    ],
                    widths=(26, 60),
                )
            ],
        ),
        Spacer(1, 6),
        RLImage(str(sketch), width=176 * mm, height=176 * 980 / 1700 * mm),
        Spacer(1, 4),
        P("Rev 2 changes: pocket opening widened 0.5 cm across all sizes; care label moved "
          "from CB neck to left side seam; carton pack changed from 20 to 24 pcs.", "small"),
        PageBreak(),
        *BUYER_HEAD(),
        P(f"Bill of materials — {O.STYLE} {O.TECHPACK_REV}", "h2"),
        P("Consumption is stated per finished piece. Wastage is the allowance to be added on "
          "top when booking; it is not included in the consumption figure.", "small"),
        Spacer(1, 4),
        grid(bom_rows, [22, 26, 76, 16, 12, 16, 12], align_right=[3, 5], font_size=7.4),
        Spacer(1, 8),
        P("Construction", "h2"),
        grid(
            [
                ["Operation", "Machine / stitch", "SPI"],
                ["Shoulder, side and sleeve seams", "overlock 4-thread, 504", "11–12"],
                ["Cuff and hem attach", "overlock 4-thread + coverstitch topstitch", "11–12"],
                ["Hood attach and CB neck tape", "overlock + single needle lockstitch 301", "12"],
                ["Zipper attach", "single needle lockstitch 301, 0.6 cm topstitch", "12"],
                ["Pocket attach", "coverstitch 602, twin needle 0.6 cm", "11"],
                ["Eyelet set", "hand press, 2 per hood, 3.5 cm either side of CF", "—"],
                ["Bartack", "hand pocket mouth ×2, zip base ×1", "—"],
            ],
            [72, 84, 24],
        ),
        Spacer(1, 8),
        two_col(
            [
                P("Care and content", "h2"),
                P(
                    "Shell: 80% cotton, 20% polyester. Rib: 95% cotton, 5% elastane.<br/>"
                    "Machine wash 30°C gentle · do not bleach · tumble dry low · "
                    "iron low, not on print · do not dry clean.<br/>"
                    "Wash before first use. Wash with similar colours.",
                    "p",
                ),
                P("Testing", "h2"),
                P(
                    "Shrinkage max 5% length / 4% width after 3 washes. Colour fastness to "
                    "washing min 4, to rubbing dry 4 / wet 3. pH 4.5–7.5. "
                    "Print crocking min 3–4. All to Nordkap RSL 2025.",
                    "p",
                ),
            ],
            [
                P("Packing", "h2"),
                P(
                    f"1 pc per polybag, folded to 300 × 420 mm. {O.PCS_PER_CARTON} pcs per "
                    "carton, solid colour solid size. Carton 600 × 400 × 350 mm, 5-ply.<br/>"
                    "Carton marking: buyer article, style, colour, size, qty, PO number, "
                    "gross and net weight, made in Bangladesh.<br/>"
                    "Hangtag with cotton string through the left cuff care label.",
                    "p",
                ),
                P("Nominated suppliers", "h2"),
                P(
                    f"Zipper: YKK, via {O.TRIMS['name']}.<br/>"
                    f"Woven and printed labels: {O.TRIMS['name']}.<br/>"
                    "Fabric: vendor's own sourcing, subject to bulk approval.",
                    "p",
                ),
            ],
        ),
    ]
    build_pdf(out / "04-tech-pack-ST-2815.pdf", st, f"Tech pack {O.STYLE}")

    bom_txt = "\n".join(
        f"{b['group']:<14}{b['ref']:<16}{b['spec'][:64]:<66}{b['cons']:>8} {b['uom']:<5}"
        f"{b['waste']:>7}   p.{b['page']}"
        for b in O.BOM
    )
    write_text(
        out / "04-tech-pack-ST-2815.paste.txt",
        f"""{O.BUYER['name']}
TECHNICAL PACKAGE — {O.STYLE} · {O.TECHPACK_REV} · {O.TECHPACK_DATE}

Style: {O.STYLE}
Buyer article: {O.BUYER_ARTICLE}
Description: {O.STYLE_DESC}
Season: {O.SEASON}
Revision: {O.TECHPACK_REV}
Issued: {O.TECHPACK_DATE}
Supersedes: Rev 1 dated 2026-08-29
Against PO: {O.PO_NO}

Rev 2 changes: pocket opening widened 0.5 cm across all sizes; care label moved from CB
neck to left side seam; carton pack changed from 20 to 24 pcs.

BILL OF MATERIALS — {O.STYLE} {O.TECHPACK_REV}
Consumption is stated per finished piece. Wastage is the allowance to be added on top when
booking; it is not included in the consumption figure.

Group         Item ref        Specification                                                  Cons/pc UoM  Wastage %  Page
{bom_txt}

CONSTRUCTION
Shoulder, side and sleeve seams — overlock 4-thread, 504 — 11-12 SPI
Cuff and hem attach — overlock 4-thread + coverstitch topstitch — 11-12 SPI
Hood attach and CB neck tape — overlock + single needle lockstitch 301 — 12 SPI
Zipper attach — single needle lockstitch 301, 0.6 cm topstitch — 12 SPI
Pocket attach — coverstitch 602, twin needle 0.6 cm — 11 SPI
Eyelet set — hand press, 2 per hood, 3.5 cm either side of CF
Bartack — hand pocket mouth x2, zip base x1

PACKING
1 pc per polybag, folded to 300 x 420 mm. {O.PCS_PER_CARTON} pcs per carton, solid colour
solid size. Carton 600 x 400 x 350 mm, 5-ply.
""",
    )
    write_json(
        out / "04-tech-pack-ST-2815.expected.json",
        {
            "_intakeKind": "tech_pack",
            "_door": "/marbim/intake → 'A tech pack'",
            "styleCode": O.STYLE,
            "lines": [
                {
                    "lineGroup": b["group"],
                    "itemRef": b["ref"],
                    "spec": b["spec"],
                    "consumption": b["cons"],
                    "uom": b["uom"],
                    "wastagePct": b["waste"],
                    "sourcePage": 2,
                }
                for b in O.BOM
            ],
            "_notes": [
                "12 lines across all four line groups. A reading that drops the packing group "
                "(it sits at the bottom of the table) is the common failure.",
                "`sourcePage` is the page of THIS pdf the table is on (2), not the 'p.4' the "
                "tech pack prints — that column refers to the buyer's own full tech pack, "
                "which the factory does not have. Either 2 or absent is right; 4/6/7/9 means "
                "the model read the wrong column.",
                "consumption is per piece and wastage is separate. 0.560 kg with 8.00% is "
                "correct; 0.6048 in consumption means the model did arithmetic it was not "
                "asked to do, and the costing studio will then double-count wastage.",
                "TRM-THR-40 is 145 m per piece — a large number with a unit that is not kg "
                "or pcs. Check it did not land as 145 pcs.",
            ],
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 05 · the amendment that breaks the L/C date (no door — typed by a human)
# ─────────────────────────────────────────────────────────────────────────────


def amendment(out: Path) -> None:
    st = [
        *BUYER_HEAD(),
        *doc_title("Purchase order amendment", f"{O.PO_NO} · {O.AMD_NO}"),
        two_col(
            [
                P("<b>Supplier</b>", "small"),
                P(O.FACTORY["name"], "b"),
                P("Attn: Merchandising / Commercial", "p"),
            ],
            [
                kv_block(
                    [
                        ("Original PO", f"{O.PO_NO} dated {O.PO_DATE}"),
                        ("Amendment", O.AMD_NO),
                        ("Amendment date", O.AMD_DATE),
                        ("Style", O.STYLE),
                    ],
                    widths=(32, 53),
                )
            ],
        ),
        Spacer(1, 9),
        P(
            "Following our floor-ready date review, the delivery window for this programme "
            "moves. Everything else on the order is unchanged.",
            "p",
        ),
        Spacer(1, 6),
        grid(
            [
                ["Field", "Was", "Now"],
                ["Ex-factory date", O.EX_FACTORY, O.AMD_EX_FACTORY],
                ["Quantity", f"{O.QTY:,} pcs", f"{O.QTY:,} pcs — unchanged"],
                ["Unit price", f"USD {money(O.UNIT_PRICE)}", f"USD {money(O.UNIT_PRICE)} — unchanged"],
                ["Colour / size grid", "as PO", "unchanged"],
                ["Shipment split", "3 shipments", "3 shipments — dates shift with ex-factory"],
            ],
            [40, 68, 72],
        ),
        Spacer(1, 9),
        P(
            f"<b>We are aware</b> that the amended ex-factory date of {O.AMD_EX_FACTORY} sits "
            f"after the latest shipment date of {O.LC_LATEST_SHIPMENT} in credit {O.LC_NO}. "
            "Our bank has been instructed to amend 44C to 2027-02-28 and 31D to 2027-03-15. "
            "<b>Do not present documents against the current credit for any shipment leaving "
            "after 2027-02-10 until the amendment reaches you through your advising bank.</b>",
            "p",
        ),
        Spacer(1, 10),
        signature_row([f"{O.BUYER['person']} · {O.BUYER['person_title']}", "Acknowledged by supplier"]),
    ]
    build_pdf(out / "05-po-amendment-AMD-01.pdf", st, f"PO amendment {O.PO_NO} {O.AMD_NO}")
    write_text(
        out / "05-po-amendment-AMD-01.paste.txt",
        f"""{O.BUYER['name']}
PURCHASE ORDER AMENDMENT — {O.PO_NO} · {O.AMD_NO}

Original PO: {O.PO_NO} dated {O.PO_DATE}
Amendment: {O.AMD_NO}
Amendment date: {O.AMD_DATE}
Style: {O.STYLE}

Field                Was                  Now
Ex-factory date      {O.EX_FACTORY}           {O.AMD_EX_FACTORY}
Quantity             {O.QTY:,} pcs           {O.QTY:,} pcs - unchanged
Unit price           USD {money(O.UNIT_PRICE)}             USD {money(O.UNIT_PRICE)} - unchanged
Colour / size grid   as PO                unchanged
Shipment split       3 shipments          3 shipments - dates shift with ex-factory

We are aware that the amended ex-factory date of {O.AMD_EX_FACTORY} sits after the latest
shipment date of {O.LC_LATEST_SHIPMENT} in credit {O.LC_NO}. Our bank has been instructed to
amend 44C to 2027-02-28 and 31D to 2027-03-15. Do not present documents against the current
credit for any shipment leaving after 2027-02-10 until the amendment reaches you through
your advising bank.
""",
    )


def build(out: Path) -> None:
    enquiry(out)
    purchase_order(out)
    breakdown_xlsx(out)
    tech_pack(out)
    amendment(out)
